# -------------------------------------------------
# excel2template.py for RDE developers
# version 1.3.2 2026/04/14
# version 1.3.1 2026/03/31
# version 1.2.1 2026/01/07

# This program is for creating dataset template definition files in RDE.
#
# Copyright (c) 2025, MDPF(Materials Data Platform), NIMS
#
# This software is released under the MIT License.
# -------------------------------------------------

__author__ = "NIMS MDPF"
__version__ = "1.3.2"
__license__ = "MIT"

from pathlib import Path
import json
from collections import defaultdict
from openpyxl import load_workbook, reader
from datetime import datetime
from dateutil import parser
import re
import argparse
import sys

reader.excel.warnings.simplefilter("ignore")


class ExcelError(Exception):
    pass


def json_dump(jdata, filepath, indent=4):
    """json形式に出力する機能"""
    print(f" - {filepath.name}を出力します。")
    with open(filepath, "w", encoding="utf_8") as f:
        json.dump(jdata, f, indent=indent, ensure_ascii=False)


def convert_value(dtype, value):
    """dtypeにあわせて値の型を変換する機能"""
    if dtype == "string":
        value = str(value)
    elif dtype == "integer":
        value = int(value)
    elif dtype == "number":
        value = float(value)
    elif dtype == "boolean":
        if value == "True":
            value = True
        else:
            value = False
    return value


def check_value(value, boolean=False):
    """値が入力されているか確認する機能"""
    if boolean:
        return (not value == "None") and (value == "True")
    else:
        if value is None:
            return False
        return not value == "None"


def get_dup_columns(d, col_name):
    """指定する列（col_name）で重複する値を返す機能"""
    vals = [x[col_name] for x in d]
    dup = set([x for x in vals if vals.count(x) > 1])
    return dup


def check_dup_params(d: dict, category_name: str, outfile: str) -> None:
    """重複するパラメータがあればエラーを出す機能"""
    dup_params = get_dup_columns(d, "parameter_name")
    if dup_params:
        raise ExcelError(
            f"要件定義（{outfile.name}）シートの{category_name=}について、重複する行が確認されました: {dup_params}"
        )


def check_req_params(d: list[dict], category_name: str, outfile: str) -> None:
    """必須項目が抜けているパラメータがあればエラーを出す機能"""
    if category_name == "metadata":
        # metadata-def
        required_keys = ["parameter_name", "name/ja", "name/en"]
    else:
        # invoice,catalog
        required_keys = ["parameter_name", "label/ja", "label/en"]

    missing_rows = []
    for i, row in enumerate(d, start=1):
        missing_keys = [key for key in required_keys if row.get(key) in (None, "None")]
        if missing_keys:
            missing_rows.append({"row": i, "missing_key": missing_keys})

    if missing_rows:
        raise ExcelError(
            f"要件定義（{outfile.name}）シートの{category_name=}について、必須項目が未指定の行が確認されました: {missing_rows}"
        )


def get_sheet_name(d):
    """key_nameとIDの対応一覧シートのシート名を取得する機能"""
    return f"{'.'.join(d[0]['key_name'].split('.')[:2])}_sample_term"


def dtype_is_expected(dtype, expected_dtypes):
    """渡された型が、渡されたパターン群に含まれるかどうかを確認する機能"""
    if dtype in expected_dtypes:
        return True
    else:
        return False


def get_validated_value(param, d, expected_dtypes, outfile):
    """JSONに格納すべき値を得る機能"""
    example = d.get("examples", None) if check_value(d.get("examples", None)) else None
    default = d.get("default", None) if check_value(d.get("default", None)) else None
    const = d.get("const") if check_value(d.get("const")) else None
    enum = d.get("enum", None).split(",") if check_value(d.get("enum", None)) else None
    required = check_value(d.get("required", None), boolean=True)
    format_v = d.get("format", None)
    pattern = d.get("pattern", None) if check_value(d.get("pattern", None)) else None
    dtype = d.get("type", None)
    sheet = outfile.name
    sheet_info = f"parameter_name={param}, {example=}, {default=}, {const=}, {sheet=}"

    # dtypeが予想される型一覧に含まれる必要あり
    if not dtype_is_expected(dtype, expected_dtypes):
        raise ExcelError(
            f"type列の値は、{'/'.join(expected_dtypes)}のいずれかとしてください。"
            f"type={dtype}, {sheet_info}"
        )

    # example列に値がない場合は、default列の値を採用する
    v = example if example else default

    # requiredがTRUEの場合は、JSONに何らかの値が格納される必要あり
    if required and not v:
        raise ExcelError(
            "required列の値がTRUEですが、JSONに格納される値がありません。"
            f"{required=}, {sheet_info}"
        )

    # const列に値がある場合は、JSONに格納される値とconst列の値とが一致している必要あり
    if const and v != const:
        raise ExcelError(f"JSONに格納される値とconst列の値が異なります。{sheet_info}")
    # enumに値がある場合は、vがenumに含まれる必要あり
    if enum and v not in enum:
        raise ExcelError(
            f"JSONに格納される値が、enumの値に含まれていません。{enum=}, {sheet_info}"
        )

    # vの型をdtypeに変更する
    if v:
        v = convert_value(dtype, v)

        # dateフォーマットに整形する
        if format_v == "date":
            # vをdatetimeオブジェクトにパースする
            date_obj = parser.parse(v)
            # datetimeオブジェクトをyyyy-mm-dd形式の文字列に変換する
            v = date_obj.strftime("%Y-%m-%d")

    # vが数値の場合、vが与えられた範囲内か調べる
    if v and dtype in ["number", "int"]:
        nmax = (
            float(d.get("maximum", None))
            if check_value(d.get("maximum", None))
            else None
        )
        exmax = (
            float(d.get("exclusiveMaximum", None))
            if check_value(d.get("exclusiveMaximum", None))
            else None
        )
        nmin = (
            float(d.get("minimum", None))
            if check_value(d.get("minimum", None))
            else None
        )
        exmin = (
            float(d.get("exclusiveMinimum", None))
            if check_value(d.get("exclusiveMinimum", None))
            else None
        )
        if (
            (nmin and v < nmin)
            or (exmin and v <= exmin)
            or (nmax and nmax < v)
            or (exmax and exmax <= v)
        ):
            raise ExcelError(
                "JSONに格納される値が指定された範囲外です。"
                f"JSONに格納される値={v},  数値上限（以上）={nmax}, 数値上限（未満）={exmax}, "
                f"数値下限（以上）={nmin}, 数値下限（より下）={exmin}, {sheet_info}"
            )

    # vが文字列の場合、文字数が与えられた範囲内か調べる
    if v and dtype == "string":
        smax = (
            int(d.get("maxLength", None))
            if check_value(d.get("maxLength", None))
            else None
        )
        smin = (
            int(d.get("minLength", None))
            if check_value(d.get("minLength", None))
            else None
        )
        slen = len(v)
        if (smin and slen < smin) or (smax and smax < slen):
            raise ExcelError(
                "JSONに格納される値が指定された範囲外です。"
                f"JSONに格納される値={v}, 最大文字数={smax}, 最小文字数={smin}, {sheet_info}"
            )

        # 正規表現での制限がある時、vが正規表現に一致するかどうかを調べる
        # pattern = "\d{4}-\d{2}-\d{2}"
        if pattern and not re.match(pattern, v):
            raise ExcelError(
                "JSONに格納される値が指定された正規表現と一致しません。"
                f"JSONに格納される値={v}, 正規表現={pattern}, {sheet_info}"
            )

    # requiredがFalse（上で判定済）で、vに何も格納されていない場合は、"null"を格納する
    v = "null" if not v else v

    return v


def read_invoice_catalog_sheet(ws):
    """invoiceとcatalogのシートからデータを取得する機能"""
    common_data = defaultdict(str)
    header = None
    data = []
    for row in ws.rows:
        # ヘッダー部が未取得の場合
        if header is None:
            if row[0].value is None:
                continue
            elif not row[0].value == "header":
                common_data[row[0].value] = str(row[1].value)
        elif row[0].value == "ヘッダー":
            continue
        # ヘッダー部を取得後
        else:
            if not row[0].value is None:
                category = row[0].value
            data.append(
                {
                    **{"category": category},
                    **{k.value: str(v.value) for k, v in zip(header, row[1:])},
                }
            )

        # ヘッダー部の取得
        if row[0].value == "header":
            header = row[1:]

    return common_data, header, data


def read_simple_sheet(ws, skipheader=0):
    """用語シートからデータを取得する機能"""
    data = []
    for row in ws.rows:
        # 不要な行はスキップする
        if str(row[0]) == "<EmptyCell>":
            continue
        # 1行目をヘッダーとする
        elif row[0].row == 1:
            header = row
        # skipheaderはスキップする
        elif row[0].row == skipheader:
            continue
        # 3行目以降は保存する
        else:
            data.append({k.value: str(v.value) for k, v in zip(header, row)})

    return data


def get_sheet(wb, sheet):
    """シートを取得する機能"""

    ws = False
    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        print(sheet + "のシートが存在しません。")

    return ws


def sheet_check(wb, output_dir, sheet):
    """対象シートの存在を確認する機能"""
    # 対象のシート名
    sheet_name = "要件定義(" + sheet + ")"

    # 対象シートがない場合はFalseを返す
    ws = get_sheet(wb, sheet_name)

    if not ws:
        return False, False

    # 対象シートがある場合はシートと出力ファイルパスを返す
    outfile = output_dir.joinpath(sheet)
    return ws, outfile


def convert_metadata_def(wb, output_dir):
    """metadata_defを出力する機能"""

    # シートのチェック
    ws, outfile = sheet_check(wb, output_dir, "metadata-def.json")

    # 対象シートがない場合は次の処理に移る
    if not ws:
        return None

    # Excelからデータを読み込む
    category_name = "metadata"
    data = read_simple_sheet(ws, skipheader=2)

    # データを抽出する
    data_on = [d for d in data if d.get("output", None) == "ON"]

    data_metadata = [d for d in data_on if d["output"] == "ON"]
    check_dup_params(data_metadata, category_name, outfile)
    check_req_params(data_metadata, category_name, outfile)

    # json形式で整理する
    jdata = defaultdict(dict)
    for d in data:
        if d.get("output", None) == "OFF":
            continue

        jdata[d.get("parameter_name", None)]["name"] = defaultdict(dict)
        jdata[d.get("parameter_name", None)]["schema"] = defaultdict(dict)

        # 項目名(日本語)
        jdata[d.get("parameter_name", None)]["name"]["ja"] = d.get("name/ja", None)
        # 項目名(英語)
        jdata[d.get("parameter_name", None)]["name"]["en"] = d.get("name/en", None)
        # データ型
        jdata[d.get("parameter_name", None)]["schema"]["type"] = d.get("type", None)
        # 表示順序
        if (
            ("order" in d)
            and d.get("order") != "None"
            and len(d.get("order").strip()) > 0
        ):
            jdata[d.get("parameter_name", None)]["order"] = int(d.get("order"))
        # フォーマット
        if check_value(d.get("format", None)):
            jdata[d.get("parameter_name", None)]["schema"]["format"] = d.get(
                "format", None
            )
        # 単位
        if check_value(d.get("unit", None)):
            jdata[d.get("parameter_name", None)]["unit"] = d.get("unit", None)
        # 説明
        if check_value(d.get("description", None)):
            jdata[d.get("parameter_name", None)]["description"] = d.get(
                "description", None
            )
        # URI
        if check_value(d.get("uri", None)):
            jdata[d.get("parameter_name", None)]["uri"] = d.get("uri", None)
        # 測定モード
        if check_value(d.get("mode", None)):
            jdata[d.get("parameter_name", None)]["mode"] = d.get("mode", None)
        # Variable)
        if check_value(d.get("variable", None)):
            jdata[d.get("parameter_name", None)]["variable"] = 1
        # 固定値
        if check_value(d.get("default", None), boolean=True):
            jdata[d.get("parameter_name", None)]["default"] = convert_value(
                d.get("type", None), d.get("sample", None)
            )
        # 装置出力
        if ("original_name" in d) and check_value(d.get("original_name", None)):
            jdata[d.get("parameter_name", None)]["original_name"] = d.get(
                "original_name", None
            )
        if ("original_type" in d) and check_value(d.get("original_type", None)):
            jdata[d.get("parameter_name", None)]["original_type"] = d.get(
                "original_type", None
            )
        if ("originalName" in d) and check_value(d.get("originalName", None)):
            jdata[d.get("parameter_name", None)]["originalName"] = d.get(
                "originalName", None
            )
        if ("originalType" in d) and check_value(d.get("originalType", None)):
            jdata[d.get("parameter_name", None)]["originalType"] = d.get(
                "originalType", None
            )

    # JSON形式で出力
    json_dump(jdata, outfile)


def _read_invoice_src_sheets(wb, output_dir, sheet_name):
    """引数で指定するシートと、2つのID対応表シートを読み込んで内容を返す機能"""

    # 一般項目の用語シートの取得
    ws_gt = get_sheet(wb, "sample.general_sample_term")

    # 分類別項目の用語シートの取得
    ws_st = get_sheet(wb, "sample.specific_sample_term")

    # 事前準備するシートがない場合は次の処理に移る
    if (not ws_st) or (not ws_gt):
        return None

    # シートのチェック
    ws, outfile = sheet_check(wb, output_dir, sheet_name)

    # 対象シートがない場合は次の処理に移る
    if not ws:
        return None

    # Excelからデータを読み込む
    common_data, header, data = read_invoice_catalog_sheet(ws)
    data_gt = read_simple_sheet(ws_gt)
    data_st = read_simple_sheet(ws_st)

    # key_nameに重複がないかチェック
    dup_keys = get_dup_columns(data_gt, "key_name")
    if dup_keys:
        sheet_name = get_sheet_name(data_gt)
        raise ExcelError(f"{sheet_name}に複数の {dup_keys}（key_name）が存在します")

    dup_keys = get_dup_columns(data_st, "key_name")
    if dup_keys:
        sheet_name = get_sheet_name(data_st)
        raise ExcelError(f"{sheet_name}に複数の {dup_keys}（key_name）が存在します")

    return common_data, data, data_gt, data_st, outfile


def _read_catalog_src_sheet(wb, output_dir, sheet_name):
    """引数で指定するシートと、2つのID対応表シートを読み込んで内容を返す機能"""

    # シートのチェック
    ws, outfile = sheet_check(wb, output_dir, sheet_name)

    # 対象シートがない場合は次の処理に移る
    if not ws:
        return None

    # Excelからデータを読み込む
    common_data, header, data = read_invoice_catalog_sheet(ws)

    return common_data, data, outfile


def _convert_invoice_schema_impl(rtn_v):
    """invoice.schema.jsonを出力する機能"""

    # 渡されたデータをそれぞれの変数に格納
    common_data, data, data_gt, data_st, outfile = rtn_v

    # json形式で整理する
    jdata = defaultdict(dict)

    # ルート部分
    # $schema
    if (common_data["$schema"] == "None") or (len(common_data["$schema"].strip()) == 0):
        raise ExcelError(f"要件定義（invoice.schema.json）の$schema、$id値は必須です。")
    jdata["$schema"] = common_data["$schema"]
    # $id
    if (common_data["$id"] == "None") or (len(common_data["$id"].strip()) == 0):
        raise ExcelError(f"要件定義（invoice.schema.json）の$schema、$id値は必須です。")
    jdata["$id"] = common_data["$id"]
    # description
    if (not common_data["description"] is None) and (
        not len(common_data["description"].strip()) == 0
    ):
        jdata["description"] = common_data["description"]
    # type
    jdata["type"] = "object"
    # required
    jdata["required"] = []
    # properties
    jdata["properties"] = defaultdict(dict)

    # customの共通部分
    if any(
        [
            d.get("output", None) != "OFF" and d.get("category", None) == "custom"
            for d in data
        ]
    ):
        jdata["required"].append("custom")

        # properties/custom
        jdata["properties"]["custom"] = defaultdict(dict)
        # properties/custom/type
        jdata["properties"]["custom"]["type"] = "object"
        # properties/custom/label/ja
        jdata["properties"]["custom"]["label"]["ja"] = "固有情報"
        # properties/custom/label/en
        jdata["properties"]["custom"]["label"]["en"] = "Custom Information"
        # properties/custom/requiredにリストを準備
        jdata["properties"]["custom"]["required"] = []
        # properties/custom/properties
        jdata["properties"]["custom"]["properties"] = defaultdict(dict)

    # sampleの共通部分
    if any(
        [
            d.get("output", None) != "OFF"
            and d.get("category", None).startswith("sample")
            for d in data
        ]
    ):
        jdata["required"].append("sample")

        # properties/sample
        jdata["properties"]["sample"] = defaultdict(dict)
        # properties/sample/type
        jdata["properties"]["sample"]["type"] = "object"
        # properties/sample/label/ja
        jdata["properties"]["sample"]["label"]["ja"] = "試料情報"
        # properties/sample/label/en
        jdata["properties"]["sample"]["label"]["en"] = "Sample Information"
        # properties/sample/properties
        jdata["properties"]["sample"]["properties"] = defaultdict(dict)

    # sample generalAttributesの部分
    if any(
        [
            d.get("output", None) != "OFF"
            and d.get("category", None) == "sample_general"
            for d in data
        ]
    ):
        # properties/sample/properties/generalAttributes
        jdata["properties"]["sample"]["properties"]["generalAttributes"] = defaultdict(
            dict
        )
        # properties/sample/properties/generalAttributes/type
        jdata["properties"]["sample"]["properties"]["generalAttributes"]["type"] = (
            "array"
        )
        # properties/sample/properties/generalAttributes/items
        jdata["properties"]["sample"]["properties"]["generalAttributes"]["items"] = []

    # sample specificAttributesの部分
    if any(
        [
            d.get("output", None) != "OFF"
            and d.get("category", None) == "sample_specific"
            for d in data
        ]
    ):
        # properties/sample/properties/specificAttributes
        jdata["properties"]["sample"]["properties"]["specificAttributes"] = defaultdict(
            dict
        )
        # properties/sample/properties/specificAttributes/type
        jdata["properties"]["sample"]["properties"]["specificAttributes"]["type"] = (
            "array"
        )
        # properties/sample/properties/specificAttributes/items
        jdata["properties"]["sample"]["properties"]["specificAttributes"]["items"] = []

    for d in data:
        if d.get("output", None) == "OFF":
            continue

        # customの部分
        if d.get("category", None) == "custom":
            jdata["properties"]["custom"]["properties"][
                d.get("parameter_name", None)
            ] = defaultdict(dict)

            # 項目名(日本語)
            jdata["properties"]["custom"]["properties"][d.get("parameter_name", None)][
                "label"
            ]["ja"] = d.get("label/ja", None)
            # 項目名(英語)
            jdata["properties"]["custom"]["properties"][d.get("parameter_name", None)][
                "label"
            ]["en"] = d.get("label/en", None)
            # データ型
            jdata["properties"]["custom"]["properties"][d.get("parameter_name", None)][
                "type"
            ] = d.get("type", None)
            # 必須項目
            if check_value(d.get("required", None), boolean=True):
                jdata["properties"]["custom"]["required"].append(
                    d.get("parameter_name", None)
                )
            # フォーマット
            if check_value(d.get("format", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["format"] = d.get("format", None)
            # 説明
            if check_value(d.get("description", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["description"] = d.get("description", None)
            # 内容サンプル
            if check_value(d.get("examples", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["examples"] = convert_value(
                    d.get("type", None), d.get("examples", None)
                )
            # 初期値
            if check_value(d.get("default", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["default"] = convert_value(
                    d.get("type", None), d.get("default", None)
                )
            # 固定値
            if check_value(d.get("const")):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["const"] = convert_value(d.get("type", None), d.get("const"))
            # 値のリスト
            if check_value(d.get("enum", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["enum"] = [
                    convert_value(d.get("type", None), v.strip())
                    for v in d.get("enum", None).split(",")
                ]
            # テキストエリア
            if check_value(d.get("options/widget", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["options"]["widget"] = d.get("options/widget", None)
            # 行数
            if check_value(d.get("options/rows", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["options"]["rows"] = int(d.get("options/rows", None))
            # 単位
            if check_value(d.get("options/unit", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["options"]["unit"] = d.get("options/unit", None)
            # プレイスホルダ
            if check_value(d.get("options/placeholder/ja", None)) or check_value(
                d.get("options/placeholder/en", None)
            ):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["options"]["placeholder"] = defaultdict(dict)
            # プレイスホルダ(日本語)
            if check_value(d.get("options/placeholder/ja", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["options"]["placeholder"]["ja"] = d.get(
                    "options/placeholder/ja", None
                )
            # プレイスホルダ(英語)
            if check_value(d.get("options/placeholder/en", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["options"]["placeholder"]["en"] = d.get(
                    "options/placeholder/en", None
                )
            # 数値上限(以下)
            if check_value(d.get("maximum", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["maximum"] = float(d.get("maximum", None))
            # 数値上限(未満)
            if check_value(d.get("exclusiveMaximum", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["exclusiveMaximum"] = float(d.get("exclusiveMaximum", None))
            # 数値下限(以上)
            if check_value(d.get("minimum", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["minimum"] = float(d.get("minimum", None))
            # 数値下限(より上)
            if check_value(d.get("exclusiveMinimum", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["exclusiveMinimum"] = float(d.get("exclusiveMinimum", None))
            # 最大文字数
            if check_value(d.get("maxLength", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["maxLength"] = int(d.get("maxLength", None))
            # 最小文字数
            if check_value(d.get("minLength", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["minLength"] = int(d.get("minLength", None))
            # 正規表現
            if check_value(d.get("pattern", None)):
                jdata["properties"]["custom"]["properties"][
                    d.get("parameter_name", None)
                ]["pattern"] = d.get("pattern", None)

        # sample_commonの部分
        if d.get("category", None) == "sample_common":
            continue

        # sample_generalの部分
        if d.get("category", None) == "sample_general":
            if d.get("output", None) == "None":
                continue
            if d.get("term", None) == "None":
                continue
            jdata["properties"]["sample"]["properties"]["generalAttributes"][
                "items"
            ].append(
                {
                    "type": "object",
                    "required": ["termId"],
                    "properties": {
                        "termId": {
                            "const": list(
                                filter(
                                    lambda x: (
                                        x["dict.term.name_ja"] == d.get("term", None)
                                    ),
                                    data_gt,
                                )
                            )[0]["term_id"]
                        }
                    },
                }
            )

        # sample_specificの部分
        if d.get("category", None) == "sample_specific":
            if d.get("output", None) == "None":
                continue
            if d.get("term", None) == "None":
                continue
            jdata["properties"]["sample"]["properties"]["specificAttributes"][
                "items"
            ].append(
                {
                    "type": "object",
                    "required": ["classId", "termId"],
                    "properties": {
                        "classId": {
                            "const": list(
                                filter(
                                    lambda x: (
                                        x["bind_class_and_term_ja"]
                                        == d.get("term", None)
                                    ),
                                    data_st,
                                )
                            )[0]["sample_class_id"]
                        },
                        "termId": {
                            "const": list(
                                filter(
                                    lambda x: (
                                        x["bind_class_and_term_ja"]
                                        == d.get("term", None)
                                    ),
                                    data_st,
                                )
                            )[0]["term_id"]
                        },
                    },
                }
            )

    # JSON形式で出力
    json_dump(jdata, outfile)


def convert_invoice_schema(wb, output_dir):
    """シートの内容を読み込み、invoice.schema.jsonを出力する機能"""

    rtn_v = _read_invoice_src_sheets(wb, output_dir, "invoice.schema.json")
    # 対象シートがない場合は次の処理に移る
    if not rtn_v:
        return None
    _convert_invoice_schema_impl(rtn_v)


def _convert_invoice_example_impl(rtn_v):
    """invoice.jsonを出力する機能"""

    expected_dtypes = ["boolean", "integer", "number", "string"]
    s = "x"
    default_uuid = f"{s * 8}-{s * 4}-{s * 4}-{s * 4}-{s * 12}"
    default_string_56 = s * 56

    # 渡されたデータをそれぞれの変数に格納
    _, data, data_gt, data_st, outfile = rtn_v

    # データを抽出する
    data_on = [d for d in data if d.get("output", None) == "ON"]

    # json形式で整理する
    jdata = defaultdict(dict)

    # basic部分
    jdata["datasetId"] = default_uuid
    jdata["basic"] = {
        "dateSubmitted": f"{datetime.today().strftime('%Y-%m-%d')}",
        "dataOwnerId": default_string_56,
        "dataName": "%%data_name%%",
        "instrumentId": default_uuid,
        "experimentId": "%%experiment_id%%",
        "description": "%%description%%",
    }

    # custom - 固有情報
    category_name = "custom"
    data_custom = [d for d in data_on if d["category"] == category_name]

    # 重複するパラメータがあればエラーを出す
    check_dup_params(data_custom, category_name, outfile)
    check_req_params(data_custom, category_name, outfile)

    for d in data_custom:
        param = d.get("parameter_name", None)
        # JSONに格納すべき値を得る
        v = get_validated_value(param, d, expected_dtypes, outfile)
        jdata["custom"][param] = v

    # sample - 資料情報
    # 資料情報全体にoutput==ONである行が１つ以上存在する場合は、sample_commonの全7行を出力する
    samples = ["sample_common", "sample_general", "sample_specific"]
    data_samples = [d for d in data_on if d.get("category", None) in samples]

    if data_samples:
        # sample_common - 資料情報（共通項目）
        category_name = "sample_common"
        data_sample_c = [
            d
            for d in data
            if check_value(d.get("parameter_name", None))
            and d.get("category", None) == category_name
        ]

        # 重複するパラメータがあればエラーを出す
        check_dup_params(data_sample_c, category_name, outfile)
        check_req_params(data_sample_c, category_name, outfile)

        # Excelのパラメータ名とJSONのプロパティ名の対応（要確認）
        param2prop = {
            "sample_name_(local_id)": "names",
            "chemical_formula_etc.": "composition",
            "administrator_(affiliation)": "ownerId",
            "reference_url": "referenceUrl",
            "related_samples": "related_samples",
            "tags": "tags",
            "description": "description",
        }

        # sampleId
        jdata["sample"]["sampleId"] = ""
        # sample_name_(local_id)のデフォルト値
        jdata["sample"][param2prop["sample_name_(local_id)"]] = []
        # administrator_(affiliation)のデフォルト値
        jdata["sample"][param2prop["administrator_(affiliation)"]] = default_string_56

        for d in data_sample_c:
            param = d.get("parameter_name", None)
            example = (
                d.get("examples", None)
                if check_value(d.get("examples", None))
                else "null"
            )

            # sample_name_(local_id)のみ、arrayとなる
            if param == "sample_name_(local_id)":
                jdata["sample"][param2prop["sample_name_(local_id)"]] = example.split(
                    ","
                )
            elif param == "administrator_(affiliation)":
                pass
            else:
                jdata["sample"][param2prop[param]] = convert_value("string", example)

        # sample_general - 資料情報（一般項目）
        category_name = "sample_general"
        data_sample_g = [d for d in data_on if d.get("category", None) == category_name]

        if data_sample_g:
            # 重複するパラメータがあればエラーを出す
            check_dup_params(data_sample_g, category_name, outfile)
            check_req_params(data_sample_g, category_name, outfile)

            generalAttributes = []

            for d in data_sample_g:
                param = d.get("parameter_name", None)
                example = (
                    d.get("examples", None)
                    if check_value(d.get("examples", None))
                    else "null"
                )
                termIds = [x["term_id"] for x in data_gt if x["key_name"] == param]
                if termIds:
                    d = {"termId": termIds[0], "value": example}
                    generalAttributes.append(d)
                else:
                    raise ExcelError(
                        f"{param}は、要件定義（{outfile.name}）シートに存在しません。"
                    )

            jdata["sample"]["generalAttributes"] = generalAttributes

        # sample_specific - 資料情報（分類別項目）
        category_name = "sample_specific"
        data_sample_s = [d for d in data_on if d.get("category", None) == category_name]

        if data_sample_s:
            # 重複するパラメータがあればエラーを出す
            check_dup_params(data_sample_s, category_name, outfile)
            check_req_params(data_sample_s, category_name, outfile)

            specificAttributes = []

            for d in data_sample_s:
                param = d.get("parameter_name", None)
                example = (
                    d.get("examples", None)
                    if check_value(d.get("examples", None))
                    else "null"
                )
                classIds = [
                    x["sample_class_id"] for x in data_st if x["key_name"] == param
                ]
                termIds = [x["term_id"] for x in data_st if x["key_name"] == param]
                if termIds:
                    d = {"classId": classIds[0], "termId": termIds[0], "value": example}
                    specificAttributes.append(d)
                else:
                    raise ExcelError(
                        f"{param}は、要件定義（{outfile.name}）シートに存在しません。"
                    )

            jdata["sample"]["specificAttributes"] = specificAttributes

    # JSON形式で出力
    outfile = outfile.parent.joinpath("invoice.json")
    json_dump(jdata, outfile, indent=2)


def convert_invoice_example(wb, output_dir):
    """シートの内容を読み込み、invoice.jsonを出力する機能"""

    rtn_v = _read_invoice_src_sheets(wb, output_dir, "invoice.schema.json")
    # 対象シートがない場合は次の処理に移る
    if not rtn_v:
        return None
    _convert_invoice_example_impl(rtn_v)


def _convert_catalog_schema_impl(rtn_v):
    """catalog.schema.jsonを出力する機能"""

    # 渡されたデータをそれぞれの変数に格納
    common_data, data, outfile = rtn_v

    # json形式で整理する
    jdata = defaultdict(dict)

    # ルート部分
    # $schema
    if (common_data["$schema"] == "None") or (len(common_data["$schema"].strip()) == 0):
        raise ExcelError(f"要件定義（catalog.schema.json）の$schema、$id値は必須です。")
    jdata["$schema"] = common_data["$schema"]
    # $id
    if (common_data["$id"] == "None") or (len(common_data["$id"].strip()) == 0):
        raise ExcelError(f"要件定義（catalog.schema.json）の$schema、$id値は必須です。")
    jdata["$id"] = common_data["$id"]

    # type
    jdata["type"] = "object"
    # required
    jdata["required"] = ["catalog"]
    # description
    if (not common_data["description"] is None) and (
        not len(common_data["description"].strip()) == 0
    ):
        jdata["description"] = common_data["description"]

    # catalog部分
    jdata["properties"]["catalog"] = defaultdict(dict)
    # properties/catalog/type
    jdata["properties"]["catalog"]["type"] = "object"
    # properties/catalog/label/ja
    jdata["properties"]["catalog"]["label"]["ja"] = common_data["title/ja"]
    # properties/catalog/label/en
    jdata["properties"]["catalog"]["label"]["en"] = common_data["title/en"]
    # properties/catalog/requiredにリストを準備
    jdata["properties"]["catalog"]["required"] = []

    # properties部分
    jdata["properties"]["catalog"]["properties"] = defaultdict(dict)

    for d in data:
        if d.get("output", None) == "OFF":
            continue

        jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)] = (
            defaultdict(dict)
        )
        # 項目名(日本語)
        jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
            "label"
        ]["ja"] = d.get("label/ja", None)
        # 項目名(英語)
        jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
            "label"
        ]["en"] = d.get("label/en", None)
        # データ型
        jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
            "type"
        ] = d.get("type", None)
        # 必須項目
        if check_value(d.get("required", None), boolean=True):
            jdata["properties"]["catalog"]["required"].append(
                d.get("parameter_name", None)
            )
        # フォーマット
        if check_value(d.get("format", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "format"
            ] = d.get("format", None)
        # 説明
        if check_value(d.get("description", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "description"
            ] = d.get("description", None)
        # 内容サンプル
        if check_value(d.get("examples", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "examples"
            ] = convert_value(d.get("type", None), d.get("examples", None))
        # 初期値
        if check_value(d.get("default", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "default"
            ] = convert_value(d.get("type", None), d.get("default", None))
        # 固定値
        if check_value(d.get("const")):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "const"
            ] = convert_value(d.get("type", None), d.get("const"))
        # 値のリスト
        if check_value(d.get("enum", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "enum"
            ] = [
                convert_value(d.get("type", None), v)
                for v in d.get("enum", None).split(",")
            ]
        # テキストエリア
        if check_value(d.get("options/widget", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "options"
            ]["widget"] = d.get("options/widget", None)
        # 行数
        if check_value(d.get("options/rows", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "options"
            ]["rows"] = int(d.get("options/rows", None))
        # 単位
        if check_value(d.get("options/unit", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "options"
            ]["unit"] = d.get("options/unit", None)
        # プレイスホルダ
        if check_value(d.get("options/placeholder/ja", None)) or check_value(
            d.get("options/placeholder/en", None)
        ):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "options"
            ]["placeholder"] = defaultdict(dict)
        # プレイスホルダ(日本語)
        if check_value(d.get("options/placeholder/ja", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "options"
            ]["placeholder"]["ja"] = d.get("options/placeholder/ja", None)
        # プレイスホルダ(英語)
        if check_value(d.get("options/placeholder/en", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "options"
            ]["placeholder"]["en"] = d.get("options/placeholder/en", None)
        # 数値上限(以下)
        if check_value(d.get("maximum", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "maximum"
            ] = float(d.get("maximum", None))
        # 数値上限(未満)
        if check_value(d.get("exclusiveMaximum", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "exclusiveMaximum"
            ] = float(d.get("exclusiveMaximum", None))
        # 数値下限(以上)
        if check_value(d.get("minimum", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "minimum"
            ] = float(d.get("minimum", None))
        # 数値下限(より上)
        if check_value(d.get("exclusiveMinimum", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "exclusiveMinimum"
            ] = float(d.get("exclusiveMinimum", None))
        # 最大文字数
        if check_value(d.get("maxLength", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "maxLength"
            ] = int(d.get("maxLength", None))
        # 最小文字数
        if check_value(d.get("minLength", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "minLength"
            ] = int(d.get("minLength", None))
        # 正規表現
        if check_value(d.get("pattern", None)):
            jdata["properties"]["catalog"]["properties"][d.get("parameter_name", None)][
                "pattern"
            ] = d.get("pattern", None)

    # JSON形式で出力
    json_dump(jdata, outfile)


def convert_catalog_schema(wb, output_dir):
    """シートの内容を読み込み、catalog.schema.jsonを出力する機能"""

    rtn_v = _read_catalog_src_sheet(wb, output_dir, "catalog.schema.json")
    # 対象シートがない場合は次の処理に移る
    if not rtn_v:
        return None
    _convert_catalog_schema_impl(rtn_v)


def _convert_catalog_example_impl(rtn_v):
    """catalog.jsonを出力する機能"""

    expected_dtypes = ["boolean", "integer", "number", "string"]
    # 渡されたデータをそれぞれの変数に格納
    common_data, data, outfile = rtn_v

    # データを抽出する
    data_on = [d for d in data if d.get("output", None) == "ON"]

    # 重複するパラメータがあればエラーを出す
    category_name = "parameter_name"
    check_dup_params(data_on, category_name, outfile)
    check_req_params(data_on, category_name, outfile)

    # json形式で整理する
    jdata = defaultdict(dict)
    jdata["$schema"] = common_data["$schema"]
    jdata["catalog"] = {}

    for d in data_on:
        param = d.get("parameter_name", None)
        v = get_validated_value(param, d, expected_dtypes, outfile)
        jdata["catalog"][param] = v

    # JSON形式で出力
    outfile = outfile.parent.joinpath("catalog.json")
    json_dump(jdata, outfile, indent=2)


def convert_catalog_example(wb, output_dir):
    """シートの内容を読み込み、catalog.jsonを出力する機能"""

    rtn_v = _read_catalog_src_sheet(wb, output_dir, "catalog.schema.json")
    # 対象シートがない場合は次の処理に移る
    if not rtn_v:
        return None
    _convert_catalog_example_impl(rtn_v)


def check_first_page(wb):
    ws = get_sheet(wb, "説明")
    nmj = ws["B1"].value
    nme = ws["B2"].value
    tid = ws["B3"].value
    if nmj is None or len(nmj) == 0:
        raise ExcelError(f"説明シートのデータセットテンプレート名は必須です。")
    if nme is None or len(nme) == 0:
        raise ExcelError(f"説明シートのデータセットテンプレート名(英)は必須です。")
    if tid is None or len(tid) == 0:
        raise ExcelError(f"説明シートのデータセットテンプレートIDは必須です。")
    return None


def main():
    parser = argparse.ArgumentParser(
        description="output some JSON files from the Excel file."
    )
    parser.add_argument(
        "--version", "-v", action="version", version="%(prog)s " + "%s" % __version__
    )
    parser.add_argument(
        "input",
        type=str,
        nargs="*",
        help="Path to the Excel file that will be the input file.",
    )
    args = parser.parse_args()

    # 入力ファイルへのパス（リスト）
    excelfiles = args.input

    # 入力ファイルが指定されていない場合は直下のExcelファイルを全て処理する
    if not excelfiles:
        excelfiles = Path.cwd().glob("*.xlsx")

    for ef in excelfiles:
        # 指定されたファイルの存在確認
        path = Path(ef)
        if not path.exists():
            print(f"指定されたファイルがありません: {path}")
            sys.exit(1)
        ef_path = Path(ef)
        print(ef_path.name + "の処理を開始します。")

        # 出力フォルダを定義して作成する
        output_dir = ef_path.parent.joinpath(ef_path.stem)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Excelファイルを開く
        wb = load_workbook(ef_path, read_only=True, data_only=True)

        # 説明ページの確認
        try:
            check_first_page(wb)
        except Exception as e:
            print(f" - 説明シートに記入漏れがありました。原因: {e}")

        # metadeta-def.jsonの出力
        try:
            convert_metadata_def(wb, output_dir)
        except Exception as e:
            print(f" - metadata-def.jsonの生成に失敗しました。原因: {e}")

        # invoice.schema.jsonの出力
        try:
            convert_invoice_schema(wb, output_dir)
        except Exception as e:
            print(f" - invoice.schema.jsonの生成に失敗しました。原因: {e}")

        # invoice.jsonの出力
        try:
            convert_invoice_example(wb, output_dir)
        except Exception as e:
            print(f" - invoice.jsonの生成に失敗しました。原因: {e}")

        # catalog.schema.jsonの出力
        try:
            convert_catalog_schema(wb, output_dir)
        except Exception as e:
            print(f" - catalog.schema.jsonの生成に失敗しました。原因: {e}")

        # catalog.jsonの出力
        try:
            convert_catalog_example(wb, output_dir)
        except Exception as e:
            print(f" - catalog.jsonの生成に失敗しました。原因: {e}")

        # Excelファイルを閉じる
        wb.close()
        print(Path(ef).name + "の処理を終了します。")
    input("Enterを押してください。")


if __name__ == "__main__":
    main()
