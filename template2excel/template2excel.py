#!/usr/bin/env python

# -------------------------------------------------
# template2excel.py
#
# This program reads an RDE template files
#   and converts it into an Excel spreadsheet.
#
# Copyright (c) 2026, MDPF(Materials Data Platform), NIMS
#
# This software is released under the MIT License.
# -------------------------------------------------

# -v(--verbosity)に関するコメント
# -vは、指定する数により表示されるメッセージが変わります
# 指定無し : 表示なし
# -v       : 期待しない値、あり得ない値を取得した場合に表示
# -vv      : 上記 + 入力、出力ディレクトリなどの指定値を表示


import argparse
import configparser
import csv
import io
import json
import os
from pathlib import Path
import sqlite3
import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font, DEFAULT_FONT
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string, get_column_letter

# from pprint import pprint

# 数式設定メモ:
# 比較的新しいExcel関数や、OpenPyXLの内部仕様に含まれていない関数を使用する場合、関数名の前に _xlfn. を付ける


# Set the new default font before creating any workbooks (全シート変更のために必要)
default_font_name = "BIZ UDPゴシック"
_font = Font(name=default_font_name)
{k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}


class ExcelError(Exception):
    pass


class CommonSettings:
    """本ツールの各種設定 (クラス変数としてのみ利用する)"""

    # 生成ツールのバージョン
    version = "ver.2026.01.20(e)"

    # セルのデフォルトの高さ
    default_height = 18.75

    # JSON Schema URL
    json_schema_url = "https://json-schema.org/draft/2020-12/schema"

    # Color code
    black = "000000"
    blue = "2F75B5"
    gray = "595959"
    green = "00BD50"
    lightgray = "D9D9D9"
    orange = "FCE4D6"
    red = "FF0000"
    white = "FFFFFF"
    # 塗りつぶし
    fill = {
        "blue": PatternFill(fgColor=f"{blue}", patternType="solid"),  # 青
        "gray": PatternFill(fgColor=f"{gray}", patternType="solid"),  # 灰色
        "green": PatternFill(fgColor=f"{green}", patternType="solid"),  # 緑
        "lightgray": PatternFill(
            fgColor=f"{lightgray}", patternType="solid"
        ),  # 薄い灰色
        "orange": PatternFill(fgColor=f"{orange}", patternType="solid"),  # オレンジ
    }
    # 条件付き書式での塗りつぶし
    pattern_fill = {
        "gray": PatternFill(
            start_color=f"{gray}", end_color=f"{gray}", fill_type="solid"
        ),
        "green": PatternFill(
            start_color=f"{green}", end_color=f"{green}", fill_type="solid"
        ),
        "orange": PatternFill(
            start_color=f"{orange}", end_color=f"{orange}", fill_type="solid"
        ),
        "red": PatternFill(start_color=f"{red}", end_color=f"{red}", fill_type="solid"),
    }
    # フォント
    font_params = {
        "name": "BIZ UDPゴシック",
    }
    font = {
        "default": Font(**font_params),
        "black": Font(color=f"{black}", **font_params),
        "lightgray": Font(color=f"{lightgray}", **font_params),
        "red": Font(color=f"{red}", **font_params),
        "white": Font(color=f"{white}", **font_params),
        "bold_black": Font(color=f"{black}", b=True, **font_params),
    }
    inline_font = {
        "bold_black": InlineFont(color=f"{black}", b=True),
        "bold_blue": InlineFont(color=f"{blue}", b=True),
        "bold_red": InlineFont(color=f"{red}", b=True),
    }
    # 罫線
    side = Side(style="thin", color=f"{black}")
    border = {
        "default": Border(top=side, bottom=side, left=side, right=side),
    }

    def __init__(self): ...


class Field:
    """各シートの一覧内の、列生成用クラス"""

    def __init__(
        self,
        name: str,
        title1: str,
        width: int = 15,
        title2: str = None,
        title2_color: str = "black",
    ):
        self.name = name
        self.title1 = title1
        self.width = width
        self.title2 = title2
        self.title2_color = title2_color

    def draw_title(self):
        """一覧の表題列を出力する機能"""
        title1 = self.title1
        title2 = self.title2
        if not title2:
            return CellRichText(title1)

        title1 = title1 + "\n"
        title2_color = self.title2_color
        match title2_color:
            case "red":
                inline_font = CommonSettings.inline_font["bold_red"]
            case "blue":
                inline_font = CommonSettings.inline_font["bold_blue"]
            case "black":
                inline_font = CommonSettings.inline_font["bold_black"]
            case _:
                # default is Black
                inline_font = CommonSettings.inline_font["bold_black"]
        return CellRichText([title1, TextBlock(inline_font, title2)])


class FieldsBase:
    """フィールド設定の基底クラス"""

    config = configparser.ConfigParser(allow_no_value=True)
    config.read_string("""

[dummy]
width = 15
title1 = ダミー
title2 = (自由記述)

""")

    def __init__(self, config: str = None):
        self.config = config if config is not None else self.__class__.config
        config = self.config
        fields = []
        f_args = {}
        for i, name in enumerate(config.sections()):
            f_args = {}

            f_args["name"] = name
            try:
                f_args["title1"] = config.get(name, "title1")
            except:
                ...
            try:
                f_args["width"] = config.get(name, "width")
            except:
                ...
            try:
                f_args["title2"] = config.get(name, "title2")
            except:
                ...
            try:
                f_args["title2_color"] = config.get(name, "title2_color")
            except:
                ...
            fields.append(Field(**f_args))
        self.fields = fields

    def list(self):
        """列名の一覧を返す ( for Debug )"""
        config = self.config
        for i, name in enumerate(config.sections()):
            print(f'name : {name} widht: {config.get(name, "width")}')

    def get_field_list(self):
        """列名のリストを返す"""
        return self.config.sections()

    def get_index_by_name(self, name) -> int:
        """指定した名前の、位置番号(0,1,2...)を返す"""
        config = self.config
        name_index = next(
            (
                i
                for i, col_name in enumerate(self.config.sections())
                if col_name == name
            ),
            -1,  # when not found
        )
        return name_index

    def get_colno_by_name(self, name) -> int:
        """カラム番号(1,2,3...)を返す"""
        name_index = self.get_index_by_name(name)
        if name_index >= 0:
            return name_index + 1
        else:
            return -1

    def get_colletter_by_name(self, name) -> str:
        """カラム名(A,B,C...)を返す"""
        col_no = self.get_colno_by_name(name)
        if col_no >= 0:
            return get_column_letter(col_no)
        else:
            return None


class FieldsMetadataDef(FieldsBase):
    """metadata-defシートの列定義"""

    config = configparser.ConfigParser(allow_no_value=True)
    config.read_string("""
[category]
width = 15
title1 = カテゴリー
title2 = (自由記述)

[output]
width = 15
title1 = 出力制御
title2 =(必ず選択)
title2_color = red

[parameter_name]
width = 65
title1 = パラメータ名
title2 = (必ず記述)
title2_color = red

[order]
width = 15
title1 = 表示順
;title2 =
;title2_color =

[original_name]
width = 30
title1 = 装置出力
title2 = (自由記述)

[original_type]
width = 30
title1 = 元のデータ型
title2 = (自由記述)

[name/ja]
width = 30 
title1 = 項目名(日本語)
title2 = (必ず記述)
title2_color = red

[name/en]
width = 55
title1 = 項目名(英語)
title2 = (必ず記述)
title2_color = red

[taxonomy]
width = 15
title1 = タクソノミー
title2 =(番号記述)

[typeformat]
width = 15
title1 = データ形式
title2 = (必ず選択)
title2_color = red

[type]
width = 15
title1 = データ型
title2 = (自動入力)

[format]
width = 15
title1 = フォーマット
title2 = (自動入力)

[unit]
width = 10
title1 = 単位
title2 = (自由記述)

[description]
width = 15
title1 = 説明
title2 = (自由記述)

[uri]
width = 20
title1 = URI
title2 = (自由記述)

[mode]
width = 10
title1 = 測定モード
title2 = (自由記述)

[variable]
width = 15
title1 = 繰り返し
title2 = (選択リスト)

[default]
width = 15
title1 = 固定値
title2 = (選択リスト)

[sample]
width = 50
title1 = サンプル

""")


class FieldsCatalogSchema(FieldsBase):
    """catalog.schemaシートの列定義"""

    config = configparser.ConfigParser(allow_no_value=True)
    config.read_string("""
[header]
width = 15
title1 = ヘッダー

[output]
width = 25
title1 = 出力制御
title2 = (必ず選択)
title2_color = red

[parameter_name]
width = 30
title1 = パラメータ名
title2 = (必ず記述)
title2_color = red

[label/ja]
width = 30
title1 = 項目名(日本語)
title2 = (必ず記述)
title2_color = red

[label/en]
width = 30
title1 = 項目名(英語)
title2 = (必ず記述)
title2_color = red

[required]
width = 15 
title1 = 必須項目
title2 = (選択リスト)

[typeformat]
width = 15
title1 = データ形式
title2 = (必ず選択)
title2_color = red

[type]
width = 15
title1 = データ型
title2 = (自動入力)

[format]
width = 15
title1 = フォーマット
title2 = (自動入力)

[options/widget]
width = 20
title1 = テキストエリア
title2 = (自動入力)

[options/rows]
width = 15
title1 = 行数
title2 = (数値記述)

[enum]
width = 20
title1 = 値のリスト
title2 = (カンマ区切り)

[description]
width = 15
title1 = 説明
title2 = (自由記述)

[examples]
width = 15
title1 = 内容サンプル
title2 = (自由記述)

[default]
width = 20
title1 = 初期値
title2 = (自由記述)

[const]
width = 20
title1 = 固定値
title2 = (自由記述)

[options/unit]
width = 30
title1 = 単位
title2 = (自由記述)

[options/placeholder/ja]
width = 30
title1 = プレイスホルダ(日本語)
title2 = (自由記述)

[options/placeholder/en]
width = 30
title1 = プレイスホルダ(英語)
title2 = (自由記述)

""")


class FieldsInvoiceSchema(FieldsBase):
    """invoice.schemaシートの列定義"""

    config = configparser.ConfigParser(allow_no_value=True)
    config.read_string("""
[header]
width = 20
title1 = ヘッダー

[category_name]
width = 20
title1 = カテゴリー名

[output]
width = 20
title1 = 出力制御
title2 = (必ず選択)
title2_color = 

[parameter_name]
width = 45
title1 = パラメータ名
title2 = (必ず記述)
title2_color = red

[term]
width = 20
title1 = 用語名
title2 = (必ず選択)
title2_color = red

[label/ja]
width = 25
title1 = 項目名(日本語)
title2 = (必ず記述)
title2_color = red

[label/en]
width = 40
title1 = 項目名(英語)
title2 = (必ず記述)
title2_color = red

[description]
width = 30
title1 = 説明
title2 = (自由記述)
title2_color = 

[examples]
width = 15
title1 = 内容サンプル
title2 = (自由記述)
title2_color = 

[options/unit]
width = 15
title1 = 単位
title2 = (自由記述)
title2_color = 

[taxonomy]
width = 15
title1 = タクソノミー
title2 = (番号記述)
title2_color = 

[required]
width = 15
title1 = 必須項目
title2 = (選択リスト)
title2_color = 

[typeformat]
width = 15
title1 = データ形式
title2 = (必ず選択)
title2_color = red

[type]
width = 10
title1 = データ型
title2 = (自動入力)
title2_color = 

[format]
width = 15
title1 = フォーマット
title2 = (自動入力)
title2_color = 

[options/widget]
width = 20
title1 = テキストエリア
title2 = (自動入力)
title2_color = 

[options/rows]
width = 15
title1 = 行数
title2 = (数値記述)
title2_color = 

[enum]
width = 15
title1 = 値のリスト
title2 = (カンマ区切り)
title2_color = 

[default]
width = 12
title1 = 初期値
title2 = (自由記述)
title2_color = 

[const]
width = 12
title1 = 固定値
title2 = (自由記述)
title2_color = 

[options/placeholder/ja]
width = 30
title1 = プレイスホルダ(日本語)
title2 = (自由記述)
title2_color = 

[options/placeholder/en]
width = 30
title1 = プレイスホルダ(英語)
title2 = (自由記述)
title2_color = 

[maximum]
width = 15
title1 = 数値上限(以下)
title2 = (数値記述)
title2_color = 

[exclusiveMaximum]
width = 22
title1 = 数値上限(未満)
title2 = (数値記述)
title2_color = 

[minimum]
width = 15
title1 = 数値下限(以上)
title2 = (数値記述)
title2_color = 

[exclusiveMinimum]
width = 22
title1 = 数値下限(より上)
title2 = (数値記述)
title2_color = 

[maxLength]
width = 15
title1 = 最大文字数
title2 = (数値記述)
title2_color = 

[minLength]
width = 15
title1 = 最小文字数
title2 = (数値記述)
title2_color = 

[pattern]
width = 15
title1 = 正規表現
title2 = (自由記述)
title2_color = 

""")


def parse_json(folder: str) -> dict:
    """指定されたフォルダのJSONファイルを読み、辞書構造にしたものを返す機能"""
    templates = {}

    p = Path(folder)
    for json_file in p.glob("*.json"):
        with open(json_file) as f:
            d = json.load(f)
        templates[json_file.name] = d

    return templates


def template_sheet(wb: Workbook) -> None:
    """各種シートのもとになるシートを作成する機能"""
    ws = wb.active
    ws.title = "template"

    for row in range(1, 1001):
        ws.row_dimensions[row].height = CommonSettings.default_height


def rm_template_sheet(wb: Workbook) -> None:
    """各種シートのもとになったシートを削除する機能"""
    sheet_to_delete = wb["template"]

    wb.remove(sheet_to_delete)


def explanation_sheet(wb: Workbook, templates: dict) -> None:
    """説明シートを作成する機能"""
    template_catalogschema = templates.get("catalog.schema.json", {})
    catalog_id = (
        template_catalogschema.get("$id", "None")
        .replace("https://rde.nims.go.jp/rde/dataset-templates/", "")
        .replace("/catalog.schema.json", "")
    )
    title_ja = template_catalogschema.get("title/ja", "None")
    title_en = template_catalogschema.get("title/en", "None")

    template_invoiceschema = templates.get("invoice.schema.json", {})
    description = template_invoiceschema.get("description", None)

    ws = wb.copy_worksheet(wb["template"])
    ws.title = "説明"

    column_width = {
        "A": 31.00,
        "B": 33.25,
        "C": 41.88,
        "F": 63.0,
    }

    for col, width in column_width.items():
        # 行の幅を変更
        ws.column_dimensions[col].width = width

    content = [
        [
            "データセットテンプレート名",
            f"{title_ja}",
            "各定義ファイルのdescriptionに利用されます",
        ],
        [
            "データセットテンプレート名(英)",
            f"{title_en}",
            "各定義ファイルのdescriptionに利用されます",
        ],
        ["データセットテンプレートID", f"{catalog_id}", "$idで利用されます"],
        ["概要", f"{description}", ""],
        ["", "", ""],
        ["作成日", "", ""],
        ["作成者", "", ""],
        ["最終更新日", "", ""],
        ["最終更新者", "", ""],
    ]

    start_row = 1
    start_col = 1  # A=1

    for i, row in enumerate(content):
        # A列
        cell = ws.cell(row=start_row + i, column=start_col)
        cell.value = row[0]
        if row[0]:
            cell.fill = CommonSettings.fill["lightgray"]
            cell.border = CommonSettings.border["default"]
        # B列
        cell = ws.cell(row=start_row + i, column=start_col + 1)
        cell.value = row[1]
        if i < 3:  # 3行目まで(->0,1,2)色つけ
            cell.fill = CommonSettings.fill["orange"]
        if row[0]:
            cell.border = CommonSettings.border["default"]
        # C列
        cell = ws.cell(row=start_row + i, column=start_col + 2)
        cell.value = row[2]
        if row[0]:
            cell.border = CommonSettings.border["default"]

    # ここまでの最終行を取得
    max_row = ws.max_row

    # シートのバージョン番号(=生成ツールのバージョン番号)
    ws[f"A{max_row + 2}"] = CommonSettings.version
    ws[f"B{max_row + 2}"] = (
        "シートレイアウトバージョン または template2excelツールのバージョン"
    )

    description = [
        ["【説明】", ""],
        ["", ""],
        ["このExcelファイルは、excel2template.exeの入力ファイルです。", ""],
        ["次の4つのシートで構成されています。", ""],
        ["", ""],
        ["・説明", ""],
        ["・要件定義(metadata-def.json)", ""],
        ["・要件定義(catalog.schema.json)", ""],
        ["・要件定義(invoice.schema.json)", ""],
        ["", ""],
        ["セルの色には以下の意味があります。", ""],
        ["", ""],
        ["", "  必須記入欄"],
        ["", "  オプション記入欄(場合によって記入)"],
        ["", "  自動記入欄(ユーザは入力不要)"],
        ["", "  記入不要欄(入力不要)"],
        ["", ""],
        ["", ""],
        ["列の削除は禁止です。", ""],
        [
            "また、必須項目が入力されていない場合はエラーになるので必ず記入をお願いします。",
            "",
        ],
    ]

    # ここまでの最終行を取得
    max_row = ws.max_row

    start_row = max_row + 2
    for i, row in enumerate(description):
        # A列
        cell = ws.cell(row=start_row + i, column=start_col)
        cell.value = row[0]
        match row[1]:
            case "  必須記入欄":
                cell.fill = CommonSettings.fill["orange"]
            case "  オプション記入欄(場合によって記入)":
                ...
            case "  自動記入欄(ユーザは入力不要)":
                cell.fill = CommonSettings.fill["blue"]
            case "  記入不要欄(入力不要)":
                cell.fill = CommonSettings.fill["gray"]
        # B列
        cell = ws.cell(row=start_row + i, column=start_col + 1)
        cell.value = row[1]

    frame = Side(style="thick", color="0000FF")

    # ここまでの最終行を取得
    max_row = ws.max_row

    for row in ws[f"A{start_row}:B{max_row}"]:
        for cell in row:
            cell.border = Border(left=frame)
    for row in ws[f"B{start_row}:B{max_row}"]:
        for cell in row:
            cell.border = Border(right=frame)
    ws[f"A{start_row}"].border = Border(left=frame, top=frame)
    ws[f"B{start_row}"].border = Border(right=frame, top=frame)
    ws[f"A{max_row}"].border = Border(left=frame, bottom=frame)
    ws[f"B{max_row}"].border = Border(right=frame, bottom=frame)


def matadata_def_sheet(wb: Workbook, template_matadatadef: dict) -> None:
    """要件定義(metadata-def.json)シートを作成する機能"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "要件定義(metadata-def.json)"

    fields = FieldsMetadataDef()
    # fields.list()
    col_list = fields.get_field_list()
    # pprint(col_list)

    start_row = 1

    for col_name in col_list:
        col_index = fields.get_index_by_name(col_name)
        col_letter = fields.get_colletter_by_name(col_name)
        field = fields.fields[col_index]
        # 列幅の設定
        ws.column_dimensions[col_letter].width = field.width
        # 項目名
        cell = ws[f"{col_letter}{start_row}"]
        cell.font = CommonSettings.font["bold_black"]
        cell.value = col_name
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        cell.fill = CommonSettings.fill["lightgray"]
        cell.border = CommonSettings.border["default"]
        # 項目名称(日本語)
        cell = ws[f"{col_letter}{start_row + 1}"]
        cell.font = CommonSettings.font["bold_black"]
        cell.value = field.draw_title()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        cell.fill = CommonSettings.fill["lightgray"]
        cell.border = CommonSettings.border["default"]

    # 表題2行目は、2行分の高さが必要
    row = start_row + 1
    current_height = ws.row_dimensions[row].height
    ws.row_dimensions[row].height = current_height * 2

    start_row = 3
    num_blank_row = 10  # 追加する余白行数

    keys = iter(template_matadatadef)
    data_length = len(template_matadatadef.keys())

    end_row = start_row + len(template_matadatadef.keys()) + num_blank_row
    for row_no in range(start_row, end_row):
        try:
            key = next(keys)
            row = template_matadatadef[key]
        except:
            # 余白行
            key = ""
            row = {}

        # category
        name = "category"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "metadata"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # parameter_name
        col_no = fields.get_colno_by_name("parameter_name")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = key
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        # output
        col_no = fields.get_colno_by_name("output")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "ON" if key else "OFF"
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = CommonSettings.border["default"]

        # order
        name = "order"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        value = row.get(name, "")
        cell.value = int(value) if value else ""
        cell.alignment = Alignment(horizontal="center")  # 数値だが中央揃え
        cell.border = CommonSettings.border["default"]

        # original_name
        name = "original_name"
        alias = "originalName"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None) or row.get(alias, None) or ""
        cell.border = CommonSettings.border["default"]

        # original_type
        name = "original_type"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, "")
        cell.border = CommonSettings.border["default"]

        # name/ja & name/en
        name = "name"
        for lang in ["ja", "en"]:
            col_no = fields.get_colno_by_name(f"{name}/{lang}")
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = row.get(name, {}).get(lang, "")
            cell.border = CommonSettings.border["default"]
            cell.fill = CommonSettings.fill["orange"]

        # taxonomy
        name = "taxonomy"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, "")
        cell.border = CommonSettings.border["default"]

        # typeformat
        name = "typeformat"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        if row:
            s_type = row.get("schema", {}).get("type", "")
            s_format = row.get("schema", {}).get("format", "")
            cell.value = _get_typeformat(s_type, s_format)
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        # type
        name = "type"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="integer","integer",'
            + f'{col_letter_of_typeformat}{row_no}="number","number",'
            + f'{col_letter_of_typeformat}{row_no}="array","array",'
            + f'{col_letter_of_typeformat}{row_no}="boolean","boolean",'
            + f'TRUE,"string")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # format
        name = "format"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'{col_letter_of_typeformat}{row_no}="datetime","date-time",'
            + f'{col_letter_of_typeformat}{row_no}="duration","duration",'
            + f'TRUE,"")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # unit
        name = "unit"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, "")
        cell.border = CommonSettings.border["default"]

        # description
        name = "description"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, "")
        cell.border = CommonSettings.border["default"]

        # uri
        name = "uri"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, "")
        cell.border = CommonSettings.border["default"]

        # mode
        name = "mode"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        value = row.get(name, "")
        cell.value = value
        cell.border = CommonSettings.border["default"]
        cell.alignment = Alignment(vertical="top", wrapText=True)
        if len(value) > 7:
            # 高さを2行分に変更
            current_height = ws.row_dimensions[row_no].height
            ws.row_dimensions[row_no].height = current_height * 2

        # variable
        name = "variable"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        value = row.get(name, "")
        cell.value = "TRUE" if value else ""
        cell.border = CommonSettings.border["default"]

        # default
        name = "default"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, "")
        cell.border = CommonSettings.border["default"]

        # sample
        name = "sample"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = ""  # 取得すべき値なし
        cell.border = CommonSettings.border["default"]

    # 先端カラム取得
    max_row = ws.max_row
    max_col = ws.max_column

    # 条件付き書式
    name = "output"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{max_row}",
        FormulaRule(
            formula=[f'{col_letter}{start_row}="ON"'],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["green"],
            font=CommonSettings.font["white"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{max_row}",
        FormulaRule(
            formula=[f'{col_letter}{start_row}="OFF"'],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
            font=CommonSettings.font["white"],
        ),
    )

    name = "output"
    col_letter_output = fields.get_colletter_by_name(name)

    # 入力必須項目
    names = [
        "parameter_name",
        "name/ja",
        "name/en",
        "typeformat",
    ]
    for name in names:
        col_letter = fields.get_colletter_by_name(name)
        ws.conditional_formatting.add(
            f"${col_letter}{start_row}:${col_letter}{max_row}",
            FormulaRule(
                formula=[
                    f'AND(${col_letter_output}{start_row}="ON", ISBLANK(${col_letter}{start_row}))'
                ],
                stopIfTrue=True,
                fill=CommonSettings.pattern_fill["red"],
            ),
        )

    # 選択肢セット
    name = "output"
    col_letter = fields.get_colletter_by_name(name)
    dv = DataValidation(
        type="list",
        formula1='"ON, OFF"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    dv.add(f"{col_letter}{start_row}:{col_letter}{max_row}")
    ws.add_data_validation(dv)

    name = "typeformat"
    col_letter = fields.get_colletter_by_name(name)
    dv = DataValidation(
        type="list",
        formula1='"string, datetime, duration, integer, number, array, boolean"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    dv.add(f"{col_letter}{start_row}:{col_letter}{max_row}")
    ws.add_data_validation(dv)


def _get_typeformat(s_type: str, s_format: str = None) -> str:
    """typeとformatの値からtypeformatの値を返す機能 (metadata-def.jsonシート用)"""
    match s_type:
        case "array":
            return s_type
        case "boolean":
            return s_type
        case "number":
            return s_type
        case "integer":
            return s_type
        case "string":
            if s_format == "date-time":
                return "datetime"  # '-'なし
            elif s_format == "duration":
                return s_format
            elif not s_format:  # formatが空の場合
                return s_type
            else:
                if args.verbosity >= 1:
                    print(
                        f"Could not get typeformat. please check type and/or format : {s_type}/{s_format}"
                    )
                return "unexpected"
        case _:
            if args.verbosity >= 1:
                print(
                    f"Could not get typeformat. please check type and/or format : {s_type}/{s_format}"
                )
            return "UNEXPECTED"

    if args.verbosity >= 1:
        print(
            f"Could not get typeformat. please check type and/or format : {s_type}/{s_format}"
        )
    return "impossible"  # ありえない


def catalog_schema_sheet(wb: Workbook, template_catalogschema: dict) -> None:
    """要件定義(catalog.schema.json)シートを作成する機能"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "要件定義(catalog.schema.json)"

    fields = FieldsCatalogSchema()
    col_list = fields.get_field_list()

    # 上部領域の作成
    ws["A1"].value = "$schema"
    ws["A2"].value = "$id"
    ws["A3"].value = "description"
    ws["A4"].value = "title/ja"
    ws["A5"].value = "title/en"

    ws["B1"].value = CommonSettings.json_schema_url
    ws["B1"].fill = CommonSettings.fill["orange"]
    ws[
        "B2"
    ].value = '="https://rde.nims.go.jp/rde/dataset-templates/"&説明!B3&"/catalog.schema.json"'
    ws["B2"].fill = CommonSettings.fill["blue"]
    ws["B2"].font = CommonSettings.font["white"]
    ws["B3"].value = '=説明!B4&":データカタログ定義"'
    ws["B3"].fill = CommonSettings.fill["blue"]
    ws["B3"].font = CommonSettings.font["white"]
    ws["B4"].value = '=説明!B1&":データカタログ定義"'
    ws["B4"].fill = CommonSettings.fill["blue"]
    ws["B4"].font = CommonSettings.font["white"]
    ws["B5"].value = '=説明!B2&":data catalog format"'
    ws["B5"].fill = CommonSettings.fill["blue"]
    ws["B5"].font = CommonSettings.font["white"]

    # セルの連結
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=1).fill = CommonSettings.fill["lightgray"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)

    max_col = ws.max_column

    # 罫線
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CommonSettings.border["default"]

    # 一覧領域の作成
    start_row = ws.max_row + 2

    for col_name in col_list:
        col_index = fields.get_index_by_name(col_name)
        col_letter = fields.get_colletter_by_name(col_name)
        field = fields.fields[col_index]
        # 列幅の設定
        ws.column_dimensions[col_letter].width = field.width
        # 項目名
        cell = ws[f"{col_letter}{start_row}"]
        cell.font = CommonSettings.font["bold_black"]
        cell.value = col_name
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        cell.fill = CommonSettings.fill["lightgray"]
        cell.border = CommonSettings.border["default"]
        # 項目名称(日本語)
        cell = ws[f"{col_letter}{start_row + 1}"]
        cell.font = CommonSettings.font["bold_black"]
        cell.value = field.draw_title()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        cell.fill = CommonSettings.fill["lightgray"]
        cell.border = CommonSettings.border["default"]

    # 表題2行目は、2行分の高さが必要
    row = start_row + 1
    current_height = ws.row_dimensions[row].height
    ws.row_dimensions[row].height = current_height * 2

    start_row = ws.max_row + 1
    num_blank_row = 10  # 追加する余白行数

    # 入力データ
    root_props = template_catalogschema.get("properties", {})

    catalog = root_props.get("catalog", {})
    required = catalog.get("required", [])

    elements = catalog.get("properties", {})

    keys = iter(elements)
    data_length = len(elements.keys())

    end_row = start_row + len(elements.keys()) + num_blank_row
    for row_no in range(start_row, end_row):
        try:
            key = next(keys)
            row = elements[key]
        except:
            # 余白行
            key = ""
            row = {}

        # header
        name = "header"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "data_catalog"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # parameter_name
        name = "parameter_name"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = key
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        # output
        name = "output"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "ON" if key else "OFF"
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = CommonSettings.border["default"]

        # label
        name = "label"
        for lang in ["ja", "en"]:
            col_no = fields.get_colno_by_name(f"{name}/{lang}")
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = row.get(name, {}).get(lang, "")
            cell.border = CommonSettings.border["default"]
            cell.fill = CommonSettings.fill["orange"]

        # required
        name = "required"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "TRUE" if key in required else ""
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = CommonSettings.border["default"]

        # typeformat
        name = "typeformat"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        if row:
            s_type = row.get("type", None)
            s_format = row.get("format", None)
            s_widget = row.get("options", {}).get("widget", None)
            s_rows = row.get("options", {}).get("rows", None)
            s_enum = row.get("enum", [])

            typeformat = _get_typeformat4catalogNinvoice(
                s_type,
                s_format,
                s_widget,
                s_enum,
            )
            cell.value = typeformat
        else:
            typeformat = None
            s_type = None
            s_format = None
            s_widget = None
            s_rows = None
            s_enum = []
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        # type
        name = "type"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="integer","integer",'
            + f'{col_letter_of_typeformat}{row_no}="number","number",'
            + f'{col_letter_of_typeformat}{row_no}="boolean","boolean",'
            + f'TRUE,"string")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # format
        name = "format"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="date","date",'
            + f'{col_letter_of_typeformat}{row_no}="markdown","markdown",'
            + f'{col_letter_of_typeformat}{row_no}="time", "time",'
            + f'{col_letter_of_typeformat}{row_no}="uri","uri",'
            + f'{col_letter_of_typeformat}{row_no}="uuid","uuid",'
            + f'TRUE,"")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # options/widget
        name = "options/widget"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="textarea","textarea",'
            + f'TRUE,"")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # options/rows
        name = "options/rows"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = s_rows if typeformat == "textarea" else ""
        cell.border = CommonSettings.border["default"]

        # enum
        name = "enum"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = ",".join(s_enum) if typeformat == "list" else ""
        cell.border = CommonSettings.border["default"]
        cell.alignment = Alignment(wrapText=True)

        # description
        name = "description"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # examples
        name = "examples"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.alignment = Alignment(wrapText=True)
        cell.border = CommonSettings.border["default"]

        # default
        name = "default"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # const
        name = "const"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # description
        name = "description"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # options/unit
        name1 = "options"
        name2 = "unit"
        col_no = fields.get_colno_by_name(f"{name1}/{name2}")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(f"{name1}", {}).get(f"{name2}", None)
        cell.border = CommonSettings.border["default"]

        # options/placeholder/{ja,en}
        name1 = "options"
        name2 = "placeholder"
        for lang in ["ja", "en"]:
            col_no = fields.get_colno_by_name(f"{name1}/{name2}/{lang}")
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = row.get(name1, {}).get(name2, {}).get(lang, "")
            cell.border = CommonSettings.border["default"]

    # 先端カラムの取得
    max_row = ws.max_row
    # max_col = ws.max_column

    # 条件付き書式
    # 上部領域
    ws.conditional_formatting.add(
        f"$B$1",
        FormulaRule(
            formula=[f"ISBLANK($B$1)"],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )
    # 一覧
    name = "output"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{max_row}",
        FormulaRule(
            formula=[f'${col_letter}{start_row}="ON"'],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["green"],
            font=CommonSettings.font["white"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{max_row}",
        FormulaRule(
            formula=[f'${col_letter}{start_row}="OFF"'],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
            font=CommonSettings.font["white"],
        ),
    )

    # 入力必須項目
    name = "output"
    col_letter_output = fields.get_colletter_by_name(name)
    name = "typeformat"
    col_letter_typeformat = fields.get_colletter_by_name(name)

    names = [
        "parameter_name",
        "label/ja",
        "label/en",
        "typeformat",
    ]
    for name in names:
        col_letter = fields.get_colletter_by_name(name)
        ws.conditional_formatting.add(
            f"${col_letter}{start_row}:${col_letter}{max_row}",
            FormulaRule(
                formula=[
                    f'AND(${col_letter_output}{start_row}="ON", ISBLANK(${col_letter}{start_row}))'
                ],
                stopIfTrue=True,
                fill=CommonSettings.pattern_fill["red"],
            ),
        )
    #
    name = "options/rows"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"${col_letter}{start_row}:${col_letter}{max_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",'
                + f'${col_letter_typeformat}{start_row}="textarea",'
                + f"ISBLANK(${col_letter}{start_row}))"
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )
    ws.conditional_formatting.add(
        f"${col_letter}{start_row}:${col_letter}{max_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",'
                + f'${col_letter_typeformat}{start_row}="textarea",'
                + f"NOT(ISBLANK(${col_letter}{start_row})))"
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["orange"],
        ),
    )
    ws.conditional_formatting.add(
        f"${col_letter}{start_row}:${col_letter}{max_row}",
        FormulaRule(
            formula=[
                f'AND({col_letter_output}{start_row}="ON", ${col_letter_typeformat}{start_row}<>"textarea")'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
        ),
    )
    #
    name = "enum"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"${col_letter}{start_row}:${col_letter}{max_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",'
                + f'${col_letter_typeformat}{start_row}="list",'
                + f"ISBLANK(${col_letter}{start_row}))"
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )
    ws.conditional_formatting.add(
        f"${col_letter}{start_row}:${col_letter}{max_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",'
                + f'${col_letter_typeformat}{start_row}="list",'
                + f"NOT(ISBLANK(${col_letter}{start_row})))"
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["orange"],
        ),
    )
    ws.conditional_formatting.add(
        f"${col_letter}{start_row}:${col_letter}{max_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON", ${col_letter_typeformat}{start_row}<>"list")'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
        ),
    )

    # 選択肢セット
    dv = DataValidation(
        type="list",
        formula1='"ON, OFF"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    dv.add(f"{col_letter_output}{start_row}:{col_letter_output}{max_row}")
    ws.add_data_validation(dv)

    dv = DataValidation(
        type="list",
        formula1='"string, textarea, number, integer, list, date, markdown, time, boolean, uri, uuid"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    dv.add(f"{col_letter_typeformat}{start_row}:{col_letter_typeformat}{max_row}")
    ws.add_data_validation(dv)

    dv = DataValidation(
        type="list",
        formula1='" ,TRUE"',  # TODO not show option of blank
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    name = "required"
    col_letter = fields.get_colletter_by_name(name)
    dv.add(f"{col_letter}{start_row}:{col_letter}{max_row}")
    ws.add_data_validation(dv)


def _get_typeformat4catalogNinvoice(
    s_type: str, s_format: str = None, s_widget: str = None, s_enum: list = []
) -> str:
    """type,format,widgetおよびenumの値からtypeformat値を返す機能"""
    match s_type:
        case "uri" | "boolean" | "integer" | "number":
            return s_type
        case "string":
            match s_format:
                case "uuid" | "date" | "time" | "markdown":
                    return s_format
            if s_enum:  # 空のリストでは"ない"とき真
                return "list"
            if s_widget == "textarea":
                return s_widget
            # その他は"string"
            return s_type
        case _:
            if args.verbosity >= 1:
                print(
                    f"Could not get typeformat. please check type,format,widget and enum : {s_type}/{s_format}/{s_widget}/{s_enum}"
                )
            return "UNEXPECTED"

    if args.verbosity >= 1:
        print(
            f"Could not get typeformat. please check type,format,widget and enum : {s_type}/{s_format}/{s_widget}/{s_enum}"
        )
    return "impossible"  # ありえない


def invoice_schema_sheet(
    wb: Workbook, template_invoiceschema: dict, conn: sqlite3.connect
) -> None:
    """要件定義(invoice.schema.json)シートを作成する機能"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "要件定義(invoice.schema.json)"

    fields = FieldsInvoiceSchema()
    col_list = fields.get_field_list()

    # 上部領域の作成
    ws["A1"].value = "$schema"
    ws["A2"].value = "$id"
    ws["A3"].value = "description"

    ws["B1"].value = CommonSettings.json_schema_url
    ws["B1"].fill = CommonSettings.fill["orange"]
    ws[
        "B2"
    ].value = '="https://rde.nims.go.jp/rde/dataset-templates/"&説明!B3&"/invoice.schema.json"'
    ws["B2"].fill = CommonSettings.fill["blue"]
    ws["B2"].font = CommonSettings.font["white"]
    ws["B3"].value = '=説明!B4&":送状定義"'
    ws["B3"].fill = CommonSettings.fill["blue"]
    ws["B3"].font = CommonSettings.font["white"]

    # セルの連結
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=1).fill = CommonSettings.fill["lightgray"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)

    max_col = ws.max_column

    # 罫線
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CommonSettings.border["default"]

    # 一覧領域の作成
    start_row = ws.max_row + 2

    for col_name in col_list:
        col_index = fields.get_index_by_name(col_name)
        col_letter = fields.get_colletter_by_name(col_name)
        field = fields.fields[col_index]
        # 列幅の設定
        ws.column_dimensions[col_letter].width = field.width
        # 項目名
        cell = ws[f"{col_letter}{start_row}"]
        cell.font = CommonSettings.font["bold_black"]
        cell.value = col_name
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        cell.fill = CommonSettings.fill["lightgray"]
        cell.border = CommonSettings.border["default"]
        # 項目名称(日本語)
        cell = ws[f"{col_letter}{start_row + 1}"]
        cell.font = CommonSettings.font["bold_black"]
        cell.value = field.draw_title()
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        cell.fill = CommonSettings.fill["lightgray"]
        cell.border = CommonSettings.border["default"]

    # 表題2行目は、2行分の高さが必要
    row = start_row + 1
    current_height = ws.row_dimensions[row].height
    ws.row_dimensions[row].height = current_height * 2

    start_row = ws.max_row + 1
    num_blank_row = 10  # 追加する余白行数

    # 入力データ
    root_props = template_invoiceschema.get("properties", {})

    custom = root_props.get("custom", {})
    props_custom = custom.get("properties", {})
    required_list = custom.get("required", [])

    sample = root_props.get("sample", {}).get("properties", {})
    sample_common = []  # dummy
    general_items = sample.get("generalAttributes", {}).get("items", [])
    specific_items = sample.get("specificAttributes", {}).get("items", [])

    # headerの種類毎の行番号(開始行と終了行の配列)を格納する変数を用意
    row_layout = {}

    # ----------------#
    # props_custom
    # ----------------#
    data_name = "props_custom"
    start_row = ws.max_row + 1
    num_blank_row = 5  # 挿入する空白行
    end_row = start_row + len(eval(data_name).keys()) + num_blank_row - 1
    row_layout[data_name] = (start_row, end_row)

    keys = iter(eval(data_name))
    for row_no in range(start_row, end_row + 1):
        try:
            key = next(keys)
            row = eval(data_name)[key]
        except:
            # 余白行
            key = ""
            row = {}

        # header
        name = "header"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "custom"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # category_name
        name = "category_name"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "固有情報"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # output
        name = "output"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "ON" if key else "OFF"
        cell.alignment = Alignment(horizontal="center")
        cell.border = CommonSettings.border["default"]

        # parameter_name
        name = "parameter_name"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        # cell.value = key
        col_letter_label_en = fields.get_colletter_by_name("label/en")
        cell.value = f'=SUBSTITUTE(LOWER(${col_letter_label_en}{row_no})," ","_")'
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # label
        name = "label"
        # enの場合の変換(1文字目大文字化、アンダースコアの空白置換、など)はしない
        for lang in ["ja", "en"]:
            col_no = fields.get_colno_by_name(f"{name}/{lang}")
            cell = ws.cell(row=row_no, column=col_no)
            value = row.get(name, {}).get(lang, None)
            if not value:
                # labelの値(とくにlabel/en)が空(または存在しない)の場合、keyで代替する
                value = key
                # warning (keyが空値でなく、ラベルとして使われた場合はお知らせ)
                if key:
                    print(
                        f'Warning: "{name}/{lang}" is not found in invoice.schema.json. Using "key"({key}) instead.'
                    )
            cell.value = value
            cell.border = CommonSettings.border["default"]
            cell.fill = CommonSettings.fill["orange"]

        # term
        name = "term"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = ""
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["gray"]

        # description
        name = "description"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # examples
        name = "examples"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        value = row.get(name, None)
        if type(value) == list:
            cell.value = ", ".join(map(str, value))
        else:
            cell.value = value
        cell.alignment = Alignment(wrapText=True)
        cell.border = CommonSettings.border["default"]

        # options/unit
        name1 = "options"
        name2 = "unit"
        col_no = fields.get_colno_by_name(f"{name1}/{name2}")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name1, {}).get(name2, None)
        cell.border = CommonSettings.border["default"]

        # taxonomy
        name = "taxonomy"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = None  # 値の取得はできない
        cell.border = CommonSettings.border["default"]

        # required
        name = "required"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "TRUE" if key in required_list else ""
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = CommonSettings.border["default"]

        dv = DataValidation(
            type="list",
            formula1='" ,TRUE"',  # TODO not show option of blank
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
        )
        dv.add(cell)
        ws.add_data_validation(dv)

        # typeformat
        name = "typeformat"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        if row:
            s_type = row.get("type", None)
            s_format = row.get("format", None)
            s_widget = row.get("options", {}).get("widget", None)
            s_rows = row.get("options", {}).get("rows", None)
            s_enum = row.get("enum", [])

            typeformat = _get_typeformat4catalogNinvoice(
                s_type,
                s_format,
                s_widget,
                s_enum,
            )
            cell.value = typeformat
        else:
            typeformat = None
            s_type = None
            s_format = None
            s_widget = None
            s_rows = None
            s_enum = []
        # cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        # type
        name = "type"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="integer","integer",'
            + f'{col_letter_of_typeformat}{row_no}="number","number",'
            + f'{col_letter_of_typeformat}{row_no}="boolean","boolean",'
            + f'TRUE,"string")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # format
        name = "format"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="date","date",'
            + f'{col_letter_of_typeformat}{row_no}="markdown","markdown",'
            + f'{col_letter_of_typeformat}{row_no}="time", "time",'
            + f'{col_letter_of_typeformat}{row_no}="uri","uri",'
            + f'{col_letter_of_typeformat}{row_no}="uuid","uuid",'
            + f'TRUE,"")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # options/widget
        name = "options/widget"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = (
            f"=_xlfn.IFS("
            + f'ISBLANK({col_letter_of_typeformat}{row_no}), "",'
            + f'{col_letter_of_typeformat}{row_no}="textarea","textarea",'
            + f'TRUE,"")'
        )
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["blue"]
        cell.font = CommonSettings.font["white"]

        # options/rows
        name = "options/rows"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = s_rows if typeformat == "textarea" else ""
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["gray"]

        # enum
        name = "enum"
        col_no = fields.get_colno_by_name(name)
        col_letter_of_typeformat = fields.get_colletter_by_name("typeformat")
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = ",".join(s_enum) if typeformat == "list" else ""
        cell.alignment = Alignment(wrapText=True)
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["gray"]

        # default
        name = "default"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # const
        name = "const"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = row.get(name, None)
        cell.border = CommonSettings.border["default"]

        # options/placeholder/{ja,en}
        name1 = "options"
        name2 = "placeholder"
        for lang in ["ja", "en"]:
            col_no = fields.get_colno_by_name(f"{name1}/{name2}/{lang}")
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = row.get(name1, {}).get(name2, {}).get(lang, "")
            cell.border = CommonSettings.border["default"]

        # max,min
        fs = (
            "maximum",
            "exclusiveMaximum",
            "minimum",
            "exclusiveMinimum",
            "maxLength",
            "minLength",
            "pattern",
        )
        for name in fs:
            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = row.get(name, None)
            cell.border = CommonSettings.border["default"]

    # ----------------#
    # sample_common (空行だけでよい)
    # ----------------#
    data_name = "sample_common"
    # pprint(ws.max_row)
    start_row = ws.max_row + 1
    # pprint(start_row)
    num_blank_row = 5  # 挿入する空白行
    end_row = start_row + len(eval(data_name)) + num_blank_row - 1
    row_layout[data_name] = (start_row, end_row)

    for row_no in range(start_row, end_row + 1):
        # header
        name = "header"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "sample_common"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # category_name
        name = "category_name"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "試料情報(共通項目)"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # output
        name = "output"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "OFF"
        cell.alignment = Alignment(horizontal="center")
        cell.border = CommonSettings.border["default"]

        # 残りのセルは、罫線のみ
        name = "parameter_name"
        start_col_no = fields.get_colno_by_name(name)
        end_col_no = ws.max_column
        for col_no in range(start_col_no, end_col_no + 1):
            cell = ws.cell(row=row_no, column=col_no)
            cell.border = CommonSettings.border["default"]

        # 背景色
        fs = ("parameter_name", "label/ja", "label/en")
        for name in fs:
            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.fill = CommonSettings.fill["orange"]
        fs = (
            "term",
            "options/unit",
            "required",
            "typeformat",
            "type",
            "format",
            "options/widget",
            "options/rows",
            "enum",
            "default",
            "const",
            "options/placeholder/ja",
            "options/placeholder/en",
            "maximum",
            "exclusiveMaximum",
            "minimum",
            "exclusiveMinimum",
            "maxLength",
            "minLength",
            "pattern",
        )
        for name in fs:
            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.fill = CommonSettings.fill["gray"]

    # ----------------#
    # general_items
    # ----------------#
    data_name = "general_items"
    start_row = ws.max_row + 1
    num_blank_row = 5  # 挿入する空白行
    end_row = start_row + len(eval(data_name)) + num_blank_row - 1
    row_layout[data_name] = (start_row, end_row)

    lines = iter(eval(data_name))
    for row_no in range(start_row, end_row + 1):
        try:
            row = next(lines)
        except:
            # 余白行
            row = {}

        term_id = row.get("properties", {}).get("termId", {}).get("const", "")
        if term_id:
            term = _get_term_w_termid(conn, term_id)
        else:
            term = None

        # header
        name = "header"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "sample_general"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # category_name
        name = "category_name"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "試料情報(一般項目)"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # output
        name = "output"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "ON" if term else "OFF"
        cell.alignment = Alignment(horizontal="center")
        cell.border = CommonSettings.border["default"]

        # term
        name = "term"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = term
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        ref_sheet = "sample.general_sample_term"
        dv = DataValidation(
            type="list",
            formula1=f"={ref_sheet}!$C$2:$C$25",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
        )
        dv.add(cell)
        ws.add_data_validation(dv)

        # etc
        col_letter_of_term = fields.get_colletter_by_name("term")
        ref_sheet = "sample.general_sample_term"

        fs = [
            ("parameter_name", 2),
            ("label/ja", 3),
            ("label/en", 4),
            ("options/placeholder/ja", 5),
            ("options/placeholder/en", 6),
        ]

        for col_info in fs:
            name = col_info[0]
            return_col_no = col_info[1]

            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = (
                f"=IFERROR("
                + f"INDEX("
                + f"{ref_sheet}!$A:$H,"
                + f"MATCH(${col_letter_of_term}{row_no},{ref_sheet}!$C:$C,FALSE),"
                + f"{return_col_no})"
                + f',"")'
            )
            # cell.border = CommonSettings.border['default']
            cell.fill = CommonSettings.fill["blue"]
            cell.font = CommonSettings.font["white"]

        # parameter_name以降のセルに、罫線設定
        name = "parameter_name"
        start_col_no = fields.get_colno_by_name(name)
        end_col_no = ws.max_column
        for col_no in range(start_col_no, end_col_no + 1):
            cell = ws.cell(row=row_no, column=col_no)
            cell.border = CommonSettings.border["default"]

        # 背景色
        fs = (
            "options/unit",
            "required",
            "typeformat",
            "type",
            "format",
            "options/widget",
            "options/rows",
            "enum",
            "default",
            "const",
            "maximum",
            "exclusiveMaximum",
            "minimum",
            "exclusiveMinimum",
            "maxLength",
            "minLength",
            "pattern",
        )
        for name in fs:
            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.fill = CommonSettings.fill["gray"]

    # ----------------#
    # specific_items
    # ----------------#
    data_name = "specific_items"
    start_row = ws.max_row + 1
    num_blank_row = 5  # 挿入する空白行
    end_row = start_row + len(eval(data_name)) + num_blank_row - 1
    row_layout[data_name] = (start_row, end_row)

    lines = iter(eval(data_name))
    for row_no in range(start_row, end_row + 1):
        try:
            row = next(lines)
        except:
            # 余白行
            row = {}

        term_id = row.get("properties", {}).get("termId", {}).get("const", "")
        class_id = row.get("properties", {}).get("classId", {}).get("const", "")

        if term_id:
            term = _get_term_w_termid_and_classid(conn, term_id, class_id)
        else:
            term = None

        # header
        name = "header"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "sample_specific"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # category_name
        name = "category_name"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "試料情報(分類別項目)"  # 固定値
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["lightgray"]

        # output
        name = "output"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = "ON" if term else "OFF"
        cell.alignment = Alignment(horizontal="center")
        cell.border = CommonSettings.border["default"]

        # term
        name = "term"
        col_no = fields.get_colno_by_name(name)
        cell = ws.cell(row=row_no, column=col_no)
        cell.value = term
        cell.border = CommonSettings.border["default"]
        cell.fill = CommonSettings.fill["orange"]

        ref_sheet = "sample.specific_sample_term"
        dv = DataValidation(
            type="list",
            formula1=f"={ref_sheet}!$L$2:$L$36",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
        )
        dv.add(cell)
        ws.add_data_validation(dv)

        # etc
        col_letter_of_term = fields.get_colletter_by_name("term")
        ref_sheet = "sample.specific_sample_term"

        fs = [
            ("parameter_name", 3),
            ("label/ja", 12),
            ("label/en", 13),
            ("options/placeholder/ja", 8),
            ("options/placeholder/en", 9),
        ]

        for col_info in fs:
            name = col_info[0]
            return_col_no = col_info[1]

            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.value = (
                f"=IFERROR("
                + f"INDEX("
                + f"{ref_sheet}!$A:$M,"
                + f"MATCH(${col_letter_of_term}{row_no},{ref_sheet}!$L:$L,FALSE),"
                + f"{return_col_no})"
                + f',"")'
            )
            cell.fill = CommonSettings.fill["blue"]
            cell.font = CommonSettings.font["white"]

        # parameter_name以降のセルに、罫線設定
        name = "parameter_name"
        start_col_no = fields.get_colno_by_name(name)
        end_col_no = ws.max_column
        for col_no in range(start_col_no, end_col_no + 1):
            cell = ws.cell(row=row_no, column=col_no)
            cell.border = CommonSettings.border["default"]

        # 背景色
        fs = (
            "options/unit",
            "required",
            "typeformat",
            "type",
            "format",
            "options/widget",
            "options/rows",
            "enum",
            "default",
            "const",
            "maximum",
            "exclusiveMaximum",
            "minimum",
            "exclusiveMinimum",
            "maxLength",
            "minLength",
            "pattern",
        )
        for name in fs:
            col_no = fields.get_colno_by_name(name)
            cell = ws.cell(row=row_no, column=col_no)
            cell.fill = CommonSettings.fill["gray"]

    # 条件付き書式
    # max_col = ws.max_column
    max_row = ws.max_row

    # 上部領域
    ws.conditional_formatting.add(
        f"$B$1",
        FormulaRule(
            formula=[f"ISBLANK($B$1)"],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )

    # 一覧

    # 全項目対象
    start_row = row_layout["props_custom"][0]
    end_row = row_layout["specific_items"][1]

    #
    name = "output"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{max_row}",
        FormulaRule(
            formula=[f'${col_letter}{start_row}="ON"'],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["green"],
            font=CommonSettings.font["white"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{max_row}",
        FormulaRule(
            formula=[f'${col_letter}{start_row}="OFF"'],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
            font=CommonSettings.font["white"],
        ),
    )
    # 項目種別
    # custom & sample_common
    start_row = row_layout["props_custom"][0]
    end_row = row_layout["sample_common"][1]

    # outputがONなら、入力必須
    name = "output"
    col_letter_output = fields.get_colletter_by_name(name)
    fs = ("label/ja", "label/en")
    for name in fs:
        col_letter = fields.get_colletter_by_name(name)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            FormulaRule(
                formula=[
                    f'AND(${col_letter_output}{start_row}="ON",ISBLANK({col_letter}{start_row}))'
                ],
                stopIfTrue=True,
                fill=CommonSettings.pattern_fill["red"],
            ),
        )
    # sample_common
    start_row = row_layout["sample_common"][0]
    end_row = row_layout["sample_common"][1]

    # outputがONなら、入力必須
    fs = ("parameter_name",)  # 1項目しかないタプルは、最後のカンマ(,)が必須
    for name in fs:
        col_letter = fields.get_colletter_by_name(name)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            FormulaRule(
                formula=[
                    f'AND(${col_letter_output}{start_row}="ON",ISBLANK({col_letter}{start_row}))'
                ],
                stopIfTrue=True,
                fill=CommonSettings.pattern_fill["red"],
            ),
        )

    # custom
    start_row = row_layout["props_custom"][0]
    end_row = row_layout["props_custom"][1]

    name = "output"
    col_letter_output = fields.get_colletter_by_name(name)

    name = "parameter_name"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",ISBLANK({col_letter}{start_row}))'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )
    name = "typeformat"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",ISBLANK({col_letter}{start_row}))'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )

    name = "typeformat"
    col_letter_typeformat = fields.get_colletter_by_name(name)

    name = "options/rows"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",${col_letter_typeformat}{start_row}="textarea",'
                + f"ISBLANK({col_letter}{start_row}))"
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",${col_letter_typeformat}{start_row}="textarea")'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["orange"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",${col_letter_typeformat}{start_row}<>"textarea")'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
        ),
    )

    name = "enum"
    col_letter = fields.get_colletter_by_name(name)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",${col_letter_typeformat}{start_row}="list",'
                + f"ISBLANK({col_letter}{start_row}))"
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["red"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",${col_letter_typeformat}{start_row}="list")'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["orange"],
        ),
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[
                f'AND(${col_letter_output}{start_row}="ON",${col_letter_typeformat}{start_row}<>"list")'
            ],
            stopIfTrue=True,
            fill=CommonSettings.pattern_fill["gray"],
        ),
    )

    # 選択肢セット (種類共通)
    start_row = row_layout["props_custom"][0]
    end_row = row_layout["specific_items"][1]

    dv = DataValidation(
        type="list",
        formula1='"ON,OFF"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    dv.add(f"{col_letter_output}{start_row}:{col_letter_output}{end_row}")
    ws.add_data_validation(dv)

    ## header と category_name  -> セルの結合にする
    sheets = ("props_custom", "sample_common", "general_items", "specific_items")
    for sheet in sheets:
        start_row = row_layout[sheet][0]
        end_row = row_layout[sheet][1]

        fs = ("header", "category_name")
        for name in fs:
            col_no = fields.get_colno_by_name(name)
            ws.merge_cells(
                start_row=start_row,
                start_column=col_no,
                end_row=end_row,
                end_column=col_no,
            )
            cell = ws.cell(row=start_row, column=col_no)
            cell.alignment = Alignment(vertical="top")


def _get_term_w_termid(conn: sqlite3.Connection, term_id: str) -> str:
    """term_idから 日本語名称を返す機能"""

    conn.row_factory = sqlite3.Row  # 戻りをlistからdict(っぽいもの)に変更
    cur1 = conn.cursor()

    tablename0 = "sample_general_sample_term"
    tablename1 = "dict_term"

    query = f"""
SELECT
     term_id
    -- ,key_name
    --,`{tablename1}`id
    ,`{tablename1}`.name_ja
  FROM
    `{tablename0}`
 INNER JOIN
    `{tablename1}`
    ON
      `{tablename0}`.term_id = `{tablename1}`.id
 WHERE
    term_id = ?
;"""

    cur1.execute(query, [term_id])
    row = cur1.fetchone()

    return row["name_ja"]


def _get_term_w_termid_and_classid(
    conn: sqlite3.Connection, term_id: str, class_id: str
) -> str:
    """term_id と class_idから その名称を返す機能"""

    conn.row_factory = sqlite3.Row  # 戻りをlistからdict(っぽいもの)に変更
    cur1 = conn.cursor()

    tablename0 = "sample_specific_sample_term"
    tablename1 = "dict_term"
    tablename2 = "sample_sample_class"

    query = f"""
SELECT
     sample_class_id
    ,term_id
    --,key_name
    ,`{tablename1}`.name_ja AS term_name
    ,`{tablename2}`.name_ja AS class_name
  FROM
    `{tablename0}`
 INNER JOIN
    `{tablename1}`
    ON
      `{tablename0}`.term_id = `{tablename1}`.id
 INNER JOIN
    `{tablename2}`
    ON
      `{tablename0}`.sample_class_id = `{tablename2}`.id
 WHERE
    term_id = ?
    AND
    sample_class_id = ?
;"""

    cur1.execute(query, [term_id, class_id])
    row = cur1.fetchone()

    return f'{row["class_name"]}/{row["term_name"]}'


def dict_term_sheet(wb: Workbook) -> None:
    """dict.termシートを作成する機能 (本シートは最終的に非表示になる)"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "dict.term"

    column_width = {
        "A": 45,
        "B": 50,
        "C": 55,
        "D": 30,
        "E": 25,
        "F": 40,
        "G": 30,
    }

    for col, width in column_width.items():
        # 行の幅を変更
        ws.column_dimensions[col].width = width

    csv_string = csv_dict_term()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    for row in reader:
        ws.append(row)

    max_col = ws.max_column
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CommonSettings.border["default"]

    # シートを非表示に変更
    ws.sheet_state = "hidden"


def sample_general_sample_term_sheet(wb: Workbook) -> None:
    """sample.general_sample_termシートを作成する機能 (本シートは最終的に非表示になる)"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "sample.general_sample_term"

    column_width = {
        "A": 40,
        "B": 50,
        "C": 25,
        "D": 40,
        "E": 45,
        "F": 45,
        "G": 50,
        "H": 40,
    }

    for col, width in column_width.items():
        # 行の幅を変更
        ws.column_dimensions[col].width = width

    csv_string = csv_sample_general_sample_term()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    for row in reader:
        ws.append(row)

    max_col = ws.max_column
    max_row = ws.max_row

    # C列 ～ H列を値から参照式に書き換え (1行目はヘッダ行なので対象外)
    #   C列 : 3
    #   H列 : max_col (rangeの右端は含まないので、+1処理が必要)
    for col in range(3, max_col + 1):
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col)
            refcol = col - 1
            cell.value = f'=VLOOKUP(A{str(row)},dict.term!A:G,{refcol},FALSE)&""'
            cell.fill = CommonSettings.fill["blue"]

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CommonSettings.border["default"]

    # シートを非表示に変更
    ws.sheet_state = "hidden"


def sample_sample_class_sheet(wb: Workbook) -> None:
    """sample.sample_classシートを作成する機能 (本シートは最終的に非表示になる)"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "sample.sample_class"

    column_width = {
        "A": 40,
        "B": 15,
        "C": 20,
    }

    for col, width in column_width.items():
        # 行の幅を変更
        ws.column_dimensions[col].width = width

    csv_string = csv_sample_sample_class()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    for row in reader:
        ws.append(row)

    max_col = ws.max_column
    max_row = ws.max_row

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CommonSettings.border["default"]

    # シートを非表示に変更
    ws.sheet_state = "hidden"


def sample_specific_sample_term_sheet(wb: Workbook) -> None:
    """sample.specific_sample_termシートを作成する機能 (本シートは最終的に非表示になる)"""
    ws = wb.copy_worksheet(wb["template"])
    ws.title = "sample.specific_sample_term"

    column_width = {
        "A": 40,
        "B": 40,
        "C": 50,
        "D": 35,
        "E": 35,
        "F": 25,
        "G": 30,
        "H": 45,
        "I": 45,
        "J": 50,
        "K": 40,
        "L": 40,
        "M": 40,
    }

    for col, width in column_width.items():
        # 行の幅を変更
        ws.column_dimensions[col].width = width

    csv_string = csv_sample_specific_sample_term()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    for row in reader:
        ws.append(row)

    max_col = ws.max_column
    max_row = ws.max_row

    # D列 ～ E列を値から参照式に書き換え (1行目はヘッダ行なので対象外)
    #   D列 : 4
    #   E列 : 5 (rangeの右端は含まないので、+1処理が必要)
    for col in range(4, 6):
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col)
            refcol = col - 2
            cell.value = (
                f'=VLOOKUP(A{str(row)},sample.sample_class!A:C,{refcol},FALSE)&""'
            )
            cell.fill = CommonSettings.fill["blue"]

    # F列 ～ K列を値から参照式に書き換え (1行目はヘッダ行なので対象外)
    #   F列 : 6
    #   K列 : 11 (rangeの右端は含まないので、+1処理が必要)
    for col in range(6, 12):
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col)
            refcol = col - 4
            cell.value = f'=VLOOKUP(B{str(row)},dict.term!A:G,{refcol},FALSE)&""'
            cell.fill = CommonSettings.fill["blue"]

    # L列 ～ M列を値から参照式に書き換え (1行目はヘッダ行なので対象外)
    #   L列 : 12
    #   M列 : 13
    col = 12
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = f'=D{str(row)}&"/"&F{str(row)}'
        cell.fill = CommonSettings.fill["blue"]
    col = 13
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = f'=E{str(row)}&"/"&G{str(row)}'
        cell.fill = CommonSettings.fill["blue"]

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CommonSettings.border["default"]

    # シートを非表示に変更
    ws.sheet_state = "hidden"


def csv_dict_term() -> str:
    """csv_dict_term のCSVデータを返す機能"""
    return """
id,name_ja,name_en,hint_ja,hint_en,term_uri,created
2f815032-c60b-2c42-46dd-cdb4624ede4b,計測装置,Characterization Instrument,,,https://matvoc.nims.go.jp/entity/Q1885,2022-07-11 04:36:42.734137+00
2f7cdff7-bc05-6fb3-3d5e-30827a252aaa,磁気共鳴,Magnetic Resonance,,,https://matvoc.nims.go.jp/entity/Q1888,2022-07-11 04:36:42.734137+00
e6bbc4b0-a89c-a7d9-65b3-08777b320816,核磁気共鳴装置,Nuclear Magnetic Resonance,,,https://matvoc.nims.go.jp/entity/Q2116,2022-07-11 04:36:42.734137+00
9cac4bf4-b310-6840-499d-59dd515b5ab5,磁気共鳴画像診断,Magnetic Resonance Imaging,,,https://matvoc.nims.go.jp/entity/Q2117,2022-07-11 04:36:42.734137+00
6268c0bb-2636-4769-6ab0-a06ae0d03b61,電子スピン共鳴,Electron Spin Resonance,,,https://matvoc.nims.go.jp/entity/Q2118,2022-07-11 04:36:42.734137+00
388d9707-6c0f-3f0b-3770-c8959e21614b,電子顕微鏡,Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1923,2022-07-11 04:36:42.734137+00
b46a3376-3789-9ed3-d3a3-bff79de11501,透過型電子顕微鏡,Transmission Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2119,2022-07-11 04:36:42.734137+00
1359fc21-f043-ccdf-7d12-33117909f55f,走査型透過電子顕微鏡,Scanning Transmission Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2120,2022-07-11 04:36:42.734137+00
ae9aa759-1af2-2155-654f-c8a2bacba3f9,走査型電子顕微鏡,Scannning Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2121,2022-07-11 04:36:42.734137+00
b4f9c6b7-8b98-bef0-3148-b65bc67cf333,超高圧電子顕微鏡,Ultra-high Voltage Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2122,2022-07-11 04:36:42.734137+00
c611ca1d-9001-8c1f-d450-2542698c3d05,クライオ電子顕微鏡,Cryo-Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2123,2022-07-11 04:36:42.734137+00
b941b1db-8b17-69c8-2129-92f48dfe26f8,三次元電子顕微鏡,3D Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2124,2022-07-11 04:36:42.734137+00
b3e00b78-ec31-5c76-15af-588847077987,光・電子相関顕微鏡,Correlative Microscopy,,,https://matvoc.nims.go.jp/entity/Q2125,2022-07-11 04:36:42.734137+00
0ae3d47d-46ea-a25a-de9d-37ca74fa3b81,光電子顕微鏡,Photoemission Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2126,2022-07-11 04:36:42.734137+00
6077793a-dc55-803d-652d-408418fb6517,低エネルギー電子顕微鏡,Low-energy Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q2127,2022-07-11 04:36:42.734137+00
7873dab9-7a2d-cf3f-bcd3-5b5341c539aa,電子線プローブマイクロアナライザー,Electron Probe Micro Analyzer,,,https://matvoc.nims.go.jp/entity/Q2128,2022-07-11 04:36:42.734137+00
918cee56-40cb-125d-0b25-697d61684bf5,試料作成・加工,Sample Prepration,,,https://matvoc.nims.go.jp/entity/Q1924,2022-07-11 04:36:42.734137+00
88bc6257-0765-f09d-b58d-8068d44c7fd7,イオンミリング,Ion Milling,,,https://matvoc.nims.go.jp/entity/Q2129,2022-07-11 04:36:42.734137+00
c91f5e74-a0ad-2cf3-48cd-da72769ad82a,集束イオンビーム,Focused Ion Beam,,,https://matvoc.nims.go.jp/entity/Q2130,2022-07-11 04:36:42.734137+00
d052fe39-bfd5-e7d2-e2a9-7435b30e0115,ウルトラミクロトーム,Ultramicrotome,,,https://matvoc.nims.go.jp/entity/Q2131,2022-07-11 04:36:42.734137+00
d7381d4f-c176-f6b5-5699-5a8861b21c85,光学顕微鏡,Optical Microscope,,,https://matvoc.nims.go.jp/entity/Q1925,2022-07-11 04:36:42.734137+00
407a48c0-4c56-c8f3-99eb-100d9048afe8,共焦点レーザー走査型顕微鏡,Confocal Laser Scanning Microscope,,,https://matvoc.nims.go.jp/entity/Q2132,2022-07-11 04:36:42.734137+00
f09a2f93-3b6f-a31f-a3eb-a3fa2dc36734,蛍光顕微鏡,Fluorescence Microscope,,,https://matvoc.nims.go.jp/entity/Q2133,2022-07-11 04:36:42.734137+00
0a44b837-a073-e2ba-0431-badf9a5cf3bc,実体顕微鏡,Stereoscopic Microscope,,,https://matvoc.nims.go.jp/entity/Q2134,2022-07-11 04:36:42.734137+00
8e5ebdad-0785-29cd-9075-0673bb3505a1,超解像顕微鏡,Super Resolution Microscope,,,https://matvoc.nims.go.jp/entity/Q2135,2022-07-11 04:36:42.734137+00
e0ea304f-deed-c4ac-010b-421eaa751ec7,位相差顕微鏡,Phase-contrast Microscope,,,https://matvoc.nims.go.jp/entity/Q2136,2022-07-11 04:36:42.734137+00
a47aa22d-5b0a-b02a-b884-7354225158e3,走査型プローブ顕微鏡,Scanning Probe Microscope,,,https://matvoc.nims.go.jp/entity/Q1926,2022-07-11 04:36:42.734137+00
b063acb0-ca84-83d0-4845-d9011c5c9cd7,走査型トンネル顕微鏡,Scanning Tunneling Microscope,,,https://matvoc.nims.go.jp/entity/Q2137,2022-07-11 04:36:42.734137+00
36c68412-4ce4-df29-c562-cdb0ca675708,原子間力顕微鏡,Atomic Force Microscope,,,https://matvoc.nims.go.jp/entity/Q2138,2022-07-11 04:36:42.734137+00
0bf7d077-2bb3-3776-fa8b-c22e3e593f9d,クロマトグラフ,Chromatograph,,,https://matvoc.nims.go.jp/entity/Q1927,2022-07-11 04:36:42.734137+00
4f9ed9a8-acb5-cb63-5c94-037529d3579a,ガスクロマトグラフ,Gas-phase Chromatograph,,,https://matvoc.nims.go.jp/entity/Q2139,2022-07-11 04:36:42.734137+00
6a29a5dc-7c02-e843-e490-b8f15509c0b4,イオンクロマトグラフ,Ion Chromatograph,,,https://matvoc.nims.go.jp/entity/Q2140,2022-07-11 04:36:42.734137+00
1d41f686-4714-663a-92c6-3ada8ba627e1,液体クロマトグラフ,Lliquid-phase Chromatograph,,,https://matvoc.nims.go.jp/entity/Q2141,2022-07-11 04:36:42.734137+00
006788db-aa32-7fb2-1e48-6498db5f2a7e,ゲル浸透クロマトグラフ,Gel Permeation Chromatograph ,,,https://matvoc.nims.go.jp/entity/Q2142,2022-07-11 04:36:42.734137+00
0db24958-a5ce-ecd0-48bd-597c4916a6e0,分光,Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1928,2022-07-11 04:36:42.734137+00
63aed514-076d-5d2c-aa3f-4dc1f50133e7,赤外分光,Infrared Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2143,2022-07-11 04:36:42.734137+00
1baaae64-280d-cb4e-8ecd-f96b2e725317,紫外・可視分光,Ultraviolet Visible Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2144,2022-07-11 04:36:42.734137+00
b1f666d2-857a-722a-29d3-c904516d319f,紫外可視近赤外分光,Ultraviolet Visible Near-Infrared Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2145,2022-07-11 04:36:42.734137+00
dc4954fe-6a80-4a2b-0ce1-daba0c6258d9,近赤外分光光度計,Nnear‐Infrared Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2146,2022-07-11 04:36:42.734137+00
727867d3-2b27-6fa8-dc2c-c5c673b27c45,蛍光分光,Flourescence Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2147,2022-07-11 04:36:42.734137+00
345bc288-717d-64eb-8219-6d91f7d9807e,誘導結合プラズマ発光分光分析計,Inductively Coupled Plasma Atomic Emission Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2148,2022-07-11 04:36:42.734137+00
1945b49e-f335-db9c-23f7-6907e2098512,X線蛍光分光分析,X-Ray Flourescence Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2149,2022-07-11 04:36:42.734137+00
518f5c41-a04f-2aeb-adf2-8a18f58bd2a0,ラマン分光,Raman Spectroscopy ,,,https://matvoc.nims.go.jp/entity/Q2150,2022-07-11 04:36:42.734137+00
4e28347c-896b-aa1e-4cf7-63efe2fa6029,円二色性分光,Circular Dichroism,,,https://matvoc.nims.go.jp/entity/Q2151,2022-07-11 04:36:42.734137+00
ebe86ea4-de1c-521e-31e9-3d4e65949a8c,X線吸収分光,X-Ray Absorption Spectroscopy ,,,https://matvoc.nims.go.jp/entity/Q2152,2022-07-11 04:36:42.734137+00
11d8693b-b689-60ea-048e-9424d9aae2b4,X線発光分光,X-Ray Emission Spectroscopy ,,,https://matvoc.nims.go.jp/entity/Q2153,2022-07-11 04:36:42.734137+00
e9a6df00-6c58-9c03-1ca2-1947701c1464,X線光電子分光,X-Ray Photoelectron  Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2154,2022-07-11 04:36:42.734137+00
43c0ebe4-0942-5198-d390-ec2ef9cee594,オージェ電子分光,Auger Electron  Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2155,2022-07-11 04:36:42.734137+00
630c766d-e6e6-4b8d-6afd-586a8d1c2d5a,光電子分光,Photoemission Electron  Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2156,2022-07-11 04:36:42.734137+00
96880150-009f-73bc-0234-d249ec2d901c,走査型X線顕微鏡,Scanning X-Ray Microscope,,,https://matvoc.nims.go.jp/entity/Q2157,2022-07-11 04:36:42.734137+00
2deb6563-25c6-09f9-502b-ff52969f0cfb,放射光,Synchrotron Radiation,,,https://matvoc.nims.go.jp/entity/Q2072,2022-07-11 04:36:42.734137+00
dc19e030-501b-0e58-ee46-8c71ecad0d03,硬X線光電子分光法,HardX-ray Photoelectron Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2158,2022-07-11 04:36:42.734137+00
aedb7f73-6b6c-a8da-db2a-86dcc260c152,装置・広域X線吸収微細構造,Extended X-ray Absorption Fine Structure,,,https://matvoc.nims.go.jp/entity/Q2159,2022-07-11 04:36:42.734137+00
34d49688-f97b-ddeb-422a-3ad4643524ea,X線吸収端近傍構造,X-ray Absorption Near Edge Structure,,,https://matvoc.nims.go.jp/entity/Q2160,2022-07-11 04:36:42.734137+00
09f70b65-778d-c5b2-efc4-f81ff4446a69,X線回折装置(放射光),X-ray Diffraction-Synchrotron Radiation,,,https://matvoc.nims.go.jp/entity/Q2161,2022-07-11 04:36:42.734137+00
ce2101b1-1ab5-68bf-e2b5-8c1c74d43703,質量分析,Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2073,2022-07-11 04:36:42.734137+00
c2d31344-4058-283e-1836-7cc6c43adc32,二重収束質量分析　,Double-Focusing Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2162,2022-07-11 04:36:42.734137+00
d0fe9e27-b791-7f92-c555-571e8d5bd634,四重極質量分析　,Quadrupole Mass Analyzer; Quadrupole Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2163,2022-07-11 04:36:42.734137+00
26a161bf-d73a-f7cc-43ab-990c6a24b509,飛行時間質量分析　,Time-Of-Flight Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2164,2022-07-11 04:36:42.734137+00
65c88d6e-8d21-1c2d-eb90-0b702a886310,イオントラップ質量分析　,Ion Trap Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2165,2022-07-11 04:36:42.734137+00
94516365-6ff5-addb-4758-dbcaafb57e2c,フーリエ変換イオンサイクロトロン共鳴質量分析　,Fourier Transfom Ion Cyclotron Resonance Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2166,2022-07-11 04:36:42.734137+00
a8be5387-4b9f-5249-525c-a80cedd6469f,飛行時間二次イオン質量分析　,Time-Of-Flight Secondary Ion Mass Spectrometry,,,https://matvoc.nims.go.jp/entity/Q2167,2022-07-11 04:36:42.734137+00
f96f9f50-ea2d-4442-23ff-d4ab8b04f49e,誘導結合プラズマ質量分析　,Inductively Coupled Plasma Mass Spectrometry,,,https://matvoc.nims.go.jp/entity/Q2168,2022-07-11 04:36:42.734137+00
cac4338b-b084-2347-7df0-2f2a0dfd4e71,マトリックス支援レーザー脱離イオン化質量分析,Maldi-Tof Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2169,2022-07-11 04:36:42.734137+00
cf32eec6-fc80-ed17-0075-0b131de1a106,二次イオン質量分析,Secondary Ion Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2170,2022-07-11 04:36:42.734137+00
522a888f-49a3-994b-f15d-c5e32d9a87ec,直接イオン化質量分析,Direct Analysis In Real Time Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2171,2022-07-11 04:36:42.734137+00
2783b277-a2e2-4053-306b-ab5c46d07f64,ガスクロマトグラフ質量分析　,Gas Chromatography - Mass Spectorometer ,,,https://matvoc.nims.go.jp/entity/Q2172,2022-07-11 04:36:42.734137+00
3c727b19-4b2e-d7e3-45c3-a41c6f7ee696,液体クロマトグラフ質量分析　,Liquid Chromatography - Mass Spectorometer ,,,https://matvoc.nims.go.jp/entity/Q2173,2022-07-11 04:36:42.734137+00
0ea8cfd2-8575-3a57-1de7-dff3b783c539,回折・散乱,Scattering & Diffraction ,,,https://matvoc.nims.go.jp/entity/Q2074,2022-07-11 04:36:42.734137+00
eed4afb5-4780-a4a2-b97a-c2135bdfad63,X線回折装置,X-Ray Diffraction,,,https://matvoc.nims.go.jp/entity/Q2263,2022-07-11 04:36:42.734137+00
387d91ec-6bf5-07d8-a688-7b3741863fff,単結晶X線回折,Single Crystal X-Ray Diffraction,,,https://matvoc.nims.go.jp/entity/Q2174,2022-07-11 04:36:42.734137+00
b99e8b5a-a4be-8c6b-2e79-963ff6d69cf8,中性子回折,Neutron Diffraction,,,https://matvoc.nims.go.jp/entity/Q2175,2022-07-11 04:36:42.734137+00
ba0f4495-a44a-f197-00b6-fe7a453f11ef,X線トポグラフィー,X-Ray Topography,,,https://matvoc.nims.go.jp/entity/Q2176,2022-07-11 04:36:42.734137+00
9c87dc8d-356e-354e-f95e-b90175cf5331,X線マイクロトモグラフィー,X-Ray Microtomography,,,https://matvoc.nims.go.jp/entity/Q2177,2022-07-11 04:36:42.734137+00
22135010-ea10-7975-4342-e9d986d90026,ラザフォード後方散乱,Rutherford Backscattering Spectrometry,,,https://matvoc.nims.go.jp/entity/Q2178,2022-07-11 04:36:42.734137+00
8a92e317-dbbd-f9f8-c6bf-1ee6979ed197,電子回折,Electron Diffraction,,,https://matvoc.nims.go.jp/entity/Q2179,2022-07-11 04:36:42.734137+00
b93499a7-88e3-c519-1cf1-c56343d946c6,磁気特性,Magnetic Characteristic,,,https://matvoc.nims.go.jp/entity/Q2075,2022-07-11 04:36:42.734137+00
f842bcb7-d862-0176-4779-ce50cecdce28,磁気特性測定システム,Magnetic Property Measurement System,,,https://matvoc.nims.go.jp/entity/Q2180,2022-07-11 04:36:42.734137+00
ebb35844-3fe5-b4e1-37b0-a728e1e4c56b,物理特性測定装置,Physical Property Measurement System,,,https://matvoc.nims.go.jp/entity/Q2181,2022-07-11 04:36:42.734137+00
e0ba0414-6e5c-0b09-cf71-dcef51e0ae2d,振動試料型磁束計,Vibrating Sample Magnetometer,,,https://matvoc.nims.go.jp/entity/Q2182,2022-07-11 04:36:42.734137+00
7cb1e759-553f-0877-e14e-19f6bf968bb0,バイオ装置,Biological,,,https://matvoc.nims.go.jp/entity/Q2076,2022-07-11 04:36:42.734137+00
92f858fb-2cc4-769b-c70e-3cf2932d471f,リアルタイムPCR装置,Real-Time PCR,,,https://matvoc.nims.go.jp/entity/Q2183,2022-07-11 04:36:42.734137+00
66afb0dd-c2ef-d420-134b-28044b99140f,PCR装置,PCR,,,https://matvoc.nims.go.jp/entity/Q2184,2022-07-11 04:36:42.734137+00
4722490b-fa74-6d8a-d0a7-d41d16978e1f,表面プラズモン共鳴装置,Surface Plasmon Resonance (SPR),,,https://matvoc.nims.go.jp/entity/Q2185,2022-07-11 04:36:42.734137+00
a15f8bd4-abe0-7e09-14eb-6a6ee6e89053,プレートリーダー,Plate Reader,,,https://matvoc.nims.go.jp/entity/Q2186,2022-07-11 04:36:42.734137+00
5455ac92-6dd7-74b7-7433-3f06a4aeac1e,レーザースキャナー,Laser Scanner,,,https://matvoc.nims.go.jp/entity/Q2187,2022-07-11 04:36:42.734137+00
3af82f29-b662-7a07-b753-83a033c828b1,フローサイトメトリー,Flow Cytometry,,,https://matvoc.nims.go.jp/entity/Q2188,2022-07-11 04:36:42.734137+00
659e8be1-701b-413c-8eb3-3988f050cddb,セルソーター,Cell Sorter,,,https://matvoc.nims.go.jp/entity/Q2189,2022-07-11 04:36:42.734137+00
f2a18d7c-65ff-67c5-982e-3413882e7db8,電気泳動装置,Electrophoresis,,,https://matvoc.nims.go.jp/entity/Q2190,2022-07-11 04:36:42.734137+00
a13bc408-1ab5-507a-aa32-4fd856823475,ゲルイメージング装置,Gel Imaging Device,,,https://matvoc.nims.go.jp/entity/Q2191,2022-07-11 04:36:42.734137+00
ff8d965d-2371-fc93-02e7-d9e5c665aabd,レーザーマイクロダイセクション,Laser Microdissection (LMD),,,https://matvoc.nims.go.jp/entity/Q2192,2022-07-11 04:36:42.734137+00
ed43d7c0-3e8f-d1bb-410c-86fc8658ad60,DNAシーケンサー,DNA Sequencer,,,https://matvoc.nims.go.jp/entity/Q2193,2022-07-11 04:36:42.734137+00
8c184331-da7c-9487-389f-3f1e00712517,その他分析装置,Analysis,,,https://matvoc.nims.go.jp/entity/Q2077,2022-07-11 04:36:42.734137+00
0f3f95ea-7ec5-4bbc-de53-5229680baefc,示差走査熱量分析,Differential Scanning Calorimetry,,,https://matvoc.nims.go.jp/entity/Q2194,2022-07-11 04:36:42.734137+00
0aadfff2-37de-411f-883a-38b62b2abbce,化学組成,Chemical composition,,,NULL,2022-10-11 06:12:34.454294+00
b2c7f83b-2d3e-b4fa-7713-c7d00be03694,熱重量分析,Thermal Gravimetric Analysis,,,https://matvoc.nims.go.jp/entity/Q2195,2022-07-11 04:36:42.734137+00
ef4d131a-caee-b971-a9de-8581650f1ebe,示差熱・熱重量同時測定,Thermal Gravimetric Differential Scanning Calorimetry,,,https://matvoc.nims.go.jp/entity/Q2196,2022-07-11 04:36:42.734137+00
88271e6e-7f12-0624-5728-93c4141649f3,熱機械分析,Thermomechanical Analyzer,,,https://matvoc.nims.go.jp/entity/Q2197,2022-07-11 04:36:42.734137+00
c809cf3c-c285-d2f7-a465-00e1db62c842,粘弾性測定,Viscoelasticity,,,https://matvoc.nims.go.jp/entity/Q2198,2022-07-11 04:36:42.734137+00
526fe0cb-d32c-152a-aa90-c02125fceec6,段差計,Profiler,,,https://matvoc.nims.go.jp/entity/Q2199,2022-07-11 04:36:42.734137+00
3cc18fdb-6113-bfac-9d4b-95bf2a2a1f0f,膜厚測定,Film Thickness Measurement,,,https://matvoc.nims.go.jp/entity/Q2081,2022-07-11 04:36:42.734137+00
8a9798a6-92d4-4eb5-c39d-f5e450d0fe4d,エリプソメーター,Ellipsometry,,,https://matvoc.nims.go.jp/entity/Q2201,2022-07-11 04:36:42.734137+00
227dc833-c1b7-eb0f-cc1a-00800ceeeb3d,接触角計,Contact Angle Meter ,,,https://matvoc.nims.go.jp/entity/Q2202,2022-07-11 04:36:42.734137+00
6ddfae24-6434-d553-c04c-0829998bd1f8,ゼータ電位,Zeta Potential,,,https://matvoc.nims.go.jp/entity/Q2203,2022-07-11 04:36:42.734137+00
c148da80-56ae-b4c5-34d0-6316a08e25e2,粒度分布測定（動的光散乱）,Dynamic Light Scattering ,,,https://matvoc.nims.go.jp/entity/Q2204,2022-07-11 04:36:42.734137+00
0245bde7-403f-2a17-4fb8-a950ac42a2a6,粒度分布測定（静的光散乱）,Static Light Scattering ,,,https://matvoc.nims.go.jp/entity/Q2205,2022-07-11 04:36:42.734137+00
bf8a0661-6c3f-3d1d-1e99-121c54031f78,蒸気圧式絶対分子量測定,Vapor Pressure Osmometer,,,https://matvoc.nims.go.jp/entity/Q2206,2022-07-11 04:36:42.734137+00
7d1c841a-5fb3-ad87-1f64-d3f1f1a5c518,電子物性評価,Electronic Property,,,https://matvoc.nims.go.jp/entity/Q2207,2022-07-11 04:36:42.734137+00
7aff20e0-f4f1-1412-acff-8cd330e7a16a,電子材料・デバイス評価,Electronic Materials & Device characterization,,,https://matvoc.nims.go.jp/entity/Q2208,2022-07-11 04:36:42.734137+00
691ba859-8db2-01b7-3b15-9dc4d041e36e,メスバウアー分光,Mössbauer Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2209,2022-07-11 04:36:42.734137+00
107226ca-1ec2-55b8-0e1f-c63743d53cf4,電気化学,Electron Chemical,,,https://matvoc.nims.go.jp/entity/Q2078,2022-07-11 04:36:42.734137+00
cf609657-03b8-7602-a54f-075f7cd2deb5,電流滴定,Amoperometry,,,https://matvoc.nims.go.jp/entity/Q2210,2022-07-11 04:36:42.734137+00
5ddb5803-ee23-f31b-6f90-40b5d610ada9,電位差測定,Potentiometry,,,https://matvoc.nims.go.jp/entity/Q2211,2022-07-11 04:36:42.734137+00
2f015dcd-3c4f-848c-cc6e-6f7b51a33e14,電流測定,Voltammetry,,,https://matvoc.nims.go.jp/entity/Q2212,2022-07-11 04:36:42.734137+00
b9314e0a-65e5-51ba-af82-a3fbbc90318a,機械特性,Mechanical Properties,,,https://matvoc.nims.go.jp/entity/Q2079,2022-07-11 04:36:42.734137+00
bc3a9fc9-506f-240c-a4f8-6214b0d4ce3d,圧縮試験,Compression Test,,,https://matvoc.nims.go.jp/entity/Q2213,2022-07-11 04:36:42.734137+00
e8bd346d-97a0-de92-9d0b-1f38e9674f53,クリープ試験,Creep Test,,,https://matvoc.nims.go.jp/entity/Q2214,2022-07-11 04:36:42.734137+00
e5265dd1-573a-bbf3-ca13-d30493c81203,動的機械分析,Dynamic Mechanical Analysis,,,https://matvoc.nims.go.jp/entity/Q2215,2022-07-11 04:36:42.734137+00
efc7807d-5204-0c30-0d99-5cdcd489ac70,疲労試験,Fatigue Testing,,,https://matvoc.nims.go.jp/entity/Q2216,2022-07-11 04:36:42.734137+00
dc7e19ad-c389-6e13-9bfd-352dd0233a1a,硬度計,Hardness Testing,,,https://matvoc.nims.go.jp/entity/Q2217,2022-07-11 04:36:42.734137+00
61b86478-b8a2-cd75-5c2e-2a271c153af2,ナノインデンテーション試験,Nanoindentation,,,https://matvoc.nims.go.jp/entity/Q2218,2022-07-11 04:36:42.734137+00
ecc91ce2-55ab-bf8f-c6c3-6c50d277291a,せん断　ねじれ,Shear or Torsion ,,,https://matvoc.nims.go.jp/entity/Q2219,2022-07-11 04:36:42.734137+00
2abe85d3-2b3a-4923-2cb8-8e7012830e90,引っ張り試験,Tension Test,,,https://matvoc.nims.go.jp/entity/Q2220,2022-07-11 04:36:42.734137+00
9f7882ce-a26f-0156-672c-dd8ff7c5eeeb,計算,Calculation,,,https://matvoc.nims.go.jp/entity/Q1921,2022-07-11 04:36:42.734137+00
9c0f62b3-cc76-dc89-3306-1dfafdbd210f,理論計算・シミュレーション,"Theory Calculation,Simulation",,,https://matvoc.nims.go.jp/entity/Q2080,2022-07-11 04:36:42.734137+00
1afbb021-8315-7848-e762-574abe266cac,理論計算, Theoritical Calculation,,,https://matvoc.nims.go.jp/entity/Q2221,2022-07-11 04:36:42.734137+00
52112a1c-05f8-8994-a1ab-eda4f2186bb7,シミュレーション,Simulation,,,https://matvoc.nims.go.jp/entity/Q2222,2022-07-11 04:36:42.734137+00
62b86c28-f65e-9033-8f4e-0b57d7d7c934,CAD,Computer-Aided Design,,,https://matvoc.nims.go.jp/entity/Q2223,2022-07-11 04:36:42.734137+00
e3e30902-355d-ac89-b710-16129b2120cd,機械学習,Machine Learning,,,https://matvoc.nims.go.jp/entity/Q2224,2022-07-11 04:36:42.734137+00
f61693c0-665e-dbaa-154d-960d1ada93ef,合成・プロセス装置,Synthesis and Processing Instruments,,,https://matvoc.nims.go.jp/entity/Q1922,2022-07-11 04:36:42.734137+00
7fd5bdad-ea92-e4ba-754b-2ed52f7deb6e,蒸着・成膜装置,"Film formation, Deposition",,,https://matvoc.nims.go.jp/entity/Q2200,2022-07-11 04:36:42.734137+00
8f1e6770-bb84-104c-02c1-96b64c9af10a,原子層堆積(ALD)装置,Atomic Layer Deposition System,,,https://matvoc.nims.go.jp/entity/Q2225,2022-07-11 04:36:42.734137+00
09decd93-eca4-8840-4768-cfd0caa5b31a,コーター,Coater,,,https://matvoc.nims.go.jp/entity/Q2226,2022-07-11 04:36:42.734137+00
32844a1b-8d3a-525b-b2a8-59e1b9f8eb9b,化学蒸着(CVD)装置,Chemical Vapor Deposition System,,,https://matvoc.nims.go.jp/entity/Q2227,2022-07-11 04:36:42.734137+00
b1443f44-9102-ac2e-896b-62f89931bf10,電着装置,Electrodeposition System,,,https://matvoc.nims.go.jp/entity/Q2228,2022-07-11 04:36:42.734137+00
20dffae5-f7b4-5bed-d32c-830e62e412bb,物理蒸着(PVD)装置,Physical Vapor Deposition System,,,https://matvoc.nims.go.jp/entity/Q2229,2022-07-11 04:36:42.734137+00
900901bf-752c-9300-4dfc-fea09eb083ca,インクジェット堆積装置,Ink-Jet Deposition System,,,https://matvoc.nims.go.jp/entity/Q2230,2022-07-11 04:36:42.734137+00
0f878cc4-61a9-3515-bbb7-bd22a118b26b,ラングミュア - ブロジェット膜堆積装置,Langmuir-Blodgett Film Deposition System,,,https://matvoc.nims.go.jp/entity/Q2231,2022-07-11 04:36:42.734137+00
04947e2e-5453-fc90-a4c9-6bf6c26863d4,プラズマ溶射装置,Plasma Spray System,,,https://matvoc.nims.go.jp/entity/Q2232,2022-07-11 04:36:42.734137+00
a32b5e67-1976-600d-fcd1-481bafd442b5,スッパタリング（スパッタ）,Sputtering,,,https://matvoc.nims.go.jp/entity/Q2233,2022-07-11 04:36:42.734137+00
035262e0-9cc6-4af5-4e87-7bb451eefa69,成形装置,"Molding,Forming",,,https://matvoc.nims.go.jp/entity/Q2082,2022-07-11 04:36:42.734137+00
517a8e5f-e8e4-7bfd-a4da-3a6fbcacecf1,冷間圧延ローラー,Cold Rollers,,,https://matvoc.nims.go.jp/entity/Q2234,2022-07-11 04:36:42.734137+00
5780fa89-658e-b898-f442-6ef06e82edb4,引抜金型,Drawing Die,,,https://matvoc.nims.go.jp/entity/Q2235,2022-07-11 04:36:42.734137+00
14183bf1-cdbc-6be8-f8a5-ad4df76d062b,押出金型,Extrusion Die,,,https://matvoc.nims.go.jp/entity/Q2236,2022-07-11 04:36:42.734137+00
e78aa7ec-19f7-922a-8f10-9322cac7842c,鍛造機械,Forging Equipment,,,https://matvoc.nims.go.jp/entity/Q2237,2022-07-11 04:36:42.734137+00
e8ff271e-ec9c-d122-0529-7aed2af0dbb7,ホットプレス,Hot Press,,,https://matvoc.nims.go.jp/entity/Q2238,2022-07-11 04:36:42.734137+00
42c1ae32-efe1-bd5b-d316-8aa804d3af7d,熱間圧延ローラー,Hot Rolling,,,https://matvoc.nims.go.jp/entity/Q2239,2022-07-11 04:36:42.734137+00
4bc5af9c-5b26-bba0-e038-e1ac1a3e5f9f,粉砕機,Mill,,,https://matvoc.nims.go.jp/entity/Q2240,2022-07-11 04:36:42.734137+00
04d740fc-90d9-37dc-e3a4-9d0e7c70dbbb,鋳型,Molding,,,https://matvoc.nims.go.jp/entity/Q2262,2022-07-11 04:36:42.734137+00
8cf7723e-dbc8-d23c-d0c5-9e94de12f98f,3Dプリンタ,3D Printer,,,https://matvoc.nims.go.jp/entity/Q2241,2022-07-11 04:36:42.734137+00
6e339ec0-31e9-0674-8a95-a7623ad41bea,リソグラフィ,Lithography,,,https://matvoc.nims.go.jp/entity/Q2083,2022-07-11 04:36:42.734137+00
f47a4970-5160-d114-691b-4e8da3509797,光露光（マスクアライナ）,Mask Aligner,,,https://matvoc.nims.go.jp/entity/Q2242,2022-07-11 04:36:42.734137+00
4f07320e-631e-6660-84b4-310edf8bdf53,光露光（ステッパ）,Stepper,,,https://matvoc.nims.go.jp/entity/Q2243,2022-07-11 04:36:42.734137+00
fa12e3c6-d3b1-7fb5-230c-e330b84a3b50,光露光（マスクレス、直接描画）,Maskless Exposure System,,,https://matvoc.nims.go.jp/entity/Q2244,2022-07-11 04:36:42.734137+00
351ff03e-330a-7c59-1053-c30417d3944e,電子線描画（EB）,Electron Beam Lithography,,,https://matvoc.nims.go.jp/entity/Q2245,2022-07-11 04:36:42.734137+00
bf99db62-9b22-8dbe-694b-263179fe42e2,ナノインプリント,Nanoimprint Lithography,,,https://matvoc.nims.go.jp/entity/Q2246,2022-07-11 04:36:42.734137+00
16eb7645-983e-a944-4664-46c838e24446,膜加工・エッチング,Etching,,,https://matvoc.nims.go.jp/entity/Q2084,2022-07-11 04:36:42.734137+00
d28cfd5c-9e50-3180-1641-076364e73e80,ドライエッチング（RIE）,Dry Etching(Reactive Ion Etching),,,https://matvoc.nims.go.jp/entity/Q2247,2022-07-11 04:36:42.734137+00
da45157e-b388-bb07-3caf-3b526ac507b4,ドライエッチング（ECR）,Dry Etching(Electron Cyclotron Resonance-RIE),,,https://matvoc.nims.go.jp/entity/Q2248,2022-07-11 04:36:42.734137+00
a96120be-a69f-8764-06a7-ec95b14a223e,ドライエッチング（その他）,Dry Etching(Others),,,https://matvoc.nims.go.jp/entity/Q2249,2022-07-11 04:36:42.734137+00
55cfd3a1-8836-7b3a-5b06-23889ebd106c,ウェット／ガスエッチング,Wet Etching/Gas Etching,,,https://matvoc.nims.go.jp/entity/Q2250,2022-07-11 04:36:42.734137+00
4d3a79b3-1287-e2ab-bc10-9b1da7432adf,レーザー加工,Laser Processing,,,https://matvoc.nims.go.jp/entity/Q2251,2022-07-11 04:36:42.734137+00
c9149367-f39d-22f1-9577-e4240a27c70f,その他加工装置,Processing,,,https://matvoc.nims.go.jp/entity/Q2085,2022-07-11 04:36:42.734137+00
ba9be651-befc-dff1-e22a-cc1d5c84a186,酸化,Oxidization System,,,https://matvoc.nims.go.jp/entity/Q2252,2022-07-11 04:36:42.734137+00
3f8f4608-d113-7ef7-a144-91174584b282,拡散,Diffusion System,,,https://matvoc.nims.go.jp/entity/Q2253,2022-07-11 04:36:42.734137+00
916c0c78-49c9-6f7b-49aa-141bf801864e,イオン注入,Ion Implantation,,,https://matvoc.nims.go.jp/entity/Q2254,2022-07-11 04:36:42.734137+00
46ebcbe1-18a5-27a5-7679-f8fc689abd94,接合,Bonder,,,https://matvoc.nims.go.jp/entity/Q2255,2022-07-11 04:36:42.734137+00
6e378247-ad78-60aa-2886-bc1cd8d0395e,レジスト塗布,Photoresist Spin Coater,,,https://matvoc.nims.go.jp/entity/Q2256,2022-07-11 04:36:42.734137+00
33b7f895-2199-99b9-b2fb-c25db239d066,現像装置,Photoresist Developer,,,https://matvoc.nims.go.jp/entity/Q2257,2022-07-11 04:36:42.734137+00
b86dfaaa-7886-7cf0-1f53-08fd7da01420,合成設備,Synthesis,,,https://matvoc.nims.go.jp/entity/Q2086,2022-07-11 04:36:42.734137+00
830cdcce-a774-adbe-2bff-5d84984aa7b5,分注機,Dispenser,,,https://matvoc.nims.go.jp/entity/Q2258,2022-07-11 04:36:42.734137+00
36a0320c-aba2-e894-553f-fb0bb3ef3f12,遠心機,Centrifuge,,,https://matvoc.nims.go.jp/entity/Q2259,2022-07-11 04:36:42.734137+00
db84df1f-4e97-af57-99e6-ae160534c851,撹拌機,Stirrer,,,https://matvoc.nims.go.jp/entity/Q2260,2022-07-11 04:36:42.734137+00
632188be-71ec-c06f-f8ba-db77e162a2e0,計測装置,Characterization Instrument,,,https://matvoc.nims.go.jp/entity/Q1884,2022-07-11 04:36:42.734137+00
ce043f2f-1bc0-1a96-f854-c8d5c633dee4,磁気共鳴,Magnetic Resonance,,,https://matvoc.nims.go.jp/entity/Q1886,2022-07-11 04:36:42.734137+00
c3612d5d-72d8-3fb2-d278-f2fc48b24f34,核磁気共鳴装置,Nuclear Magnetic Resonance,,,https://matvoc.nims.go.jp/entity/Q2108,2022-07-11 04:36:42.734137+00
a154317d-f57a-ab21-0deb-dcec47e92b31,磁気共鳴画像診断,Magnetic Resonance Imaging,,,https://matvoc.nims.go.jp/entity/Q1933,2022-07-11 04:36:42.734137+00
2fa6eb12-2513-85b0-b4eb-18d655381120,電子スピン共鳴,Electron Spin Resonance,,,https://matvoc.nims.go.jp/entity/Q1929,2022-07-11 04:36:42.734137+00
69b1984e-09b1-6f51-39f7-1cc828cdbe6f,電子顕微鏡,Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1889,2022-07-11 04:36:42.734137+00
6d66f347-5d24-0714-7f7c-ce00ea124a9a,透過型電子顕微鏡,Transmission Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1934,2022-07-11 04:36:42.734137+00
9ac2c04c-e21f-4543-3c97-0a962d566460,走査型透過電子顕微鏡,Scanning Transmission Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1935,2022-07-11 04:36:42.734137+00
eefd0d7a-3bc9-2796-679c-ecc6f5d933f1,走査型電子顕微鏡,Scannning Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1936,2022-07-11 04:36:42.734137+00
cf621da8-59d1-f2a9-66a5-f1bf38613e0e,超高圧電子顕微鏡,Ultra-high Voltage Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1937,2022-07-11 04:36:42.734137+00
7e882eec-11b0-e695-31ef-de2146c66375,クライオ電子顕微鏡,Cryo-Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1938,2022-07-11 04:36:42.734137+00
61c48cf4-3ec6-2e83-6680-fb83f800c306,三次元電子顕微鏡,3D Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1939,2022-07-11 04:36:42.734137+00
1bc80a36-7caa-f24b-7264-0dbcdf24e383,光・電子相関顕微鏡,Correlative Microscopy,,,https://matvoc.nims.go.jp/entity/Q1940,2022-07-11 04:36:42.734137+00
3e1f1ec0-6610-3e79-846f-68120dca2758,光電子顕微鏡,Photoemission Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1941,2022-07-11 04:36:42.734137+00
090d6535-655d-3e4a-da78-e80dca0bd0b5,低エネルギー電子顕微鏡,Low-energy Electron Microscope,,,https://matvoc.nims.go.jp/entity/Q1942,2022-07-11 04:36:42.734137+00
9ea607e7-01e2-3595-9aa5-5c443ead9572,電子線プローブマイクロアナライザー,Electron Probe Micro Analyzer,,,https://matvoc.nims.go.jp/entity/Q1930,2022-07-11 04:36:42.734137+00
4e6a790c-5ced-53e9-183a-ec0f869d2a84,試料作成・加工,Sample Prepration,,,https://matvoc.nims.go.jp/entity/Q1890,2022-07-11 04:36:42.734137+00
2000a06d-8374-6660-acdb-7e5bb0856a85,イオンミリング,Ion Milling,,,https://matvoc.nims.go.jp/entity/Q1943,2022-07-11 04:36:42.734137+00
1c4526fa-7569-3baf-24dc-2ef00c5cfb77,集束イオンビーム,Focused Ion Beam,,,https://matvoc.nims.go.jp/entity/Q1944,2022-07-11 04:36:42.734137+00
d901e594-40a8-e7a8-9c0e-9c1225662224,ウルトラミクロトーム,Ultramicrotome,,,https://matvoc.nims.go.jp/entity/Q1931,2022-07-11 04:36:42.734137+00
760a79e1-54b1-c6ad-2fdf-87fe621bce35,光学顕微鏡,Optical Microscope,,,https://matvoc.nims.go.jp/entity/Q1891,2022-07-11 04:36:42.734137+00
e55af366-adff-cbb0-8043-19503b20efb8,共焦点レーザー走査型顕微鏡,Confocal Laser Scanning Microscope,,,https://matvoc.nims.go.jp/entity/Q1945,2022-07-11 04:36:42.734137+00
388c2e56-3a10-8097-00cd-722408cb2288,蛍光顕微鏡,Fluorescence Microscope,,,https://matvoc.nims.go.jp/entity/Q1946,2022-07-11 04:36:42.734137+00
5e166ac4-bfcd-457a-84bc-8626abe9188f,購入元,Supplier,,,NULL,2022-10-11 06:12:52.876335+00
e854d8d4-a9ce-a6c2-fc35-aaa029d3ceee,実体顕微鏡,Stereoscopic Microscope,,,https://matvoc.nims.go.jp/entity/Q1947,2022-07-11 04:36:42.734137+00
b7159996-c587-8b29-da03-e42f12361d33,超解像顕微鏡,Super Resolution Microscope,,,https://matvoc.nims.go.jp/entity/Q1948,2022-07-11 04:36:42.734137+00
b8492720-a7e7-eec9-1a1a-bf07c0533620,位相差顕微鏡,Phase-contrast Microscope,,,https://matvoc.nims.go.jp/entity/Q1932,2022-07-11 04:36:42.734137+00
61380b4d-a946-e0bd-0f47-41d8e1a869a4,走査型プローブ顕微鏡,Scanning Probe Microscope,,,https://matvoc.nims.go.jp/entity/Q1892,2022-07-11 04:36:42.734137+00
ef9f8865-05ff-1b46-3a39-bc0a14f9a0ce,走査型トンネル顕微鏡,Scanning Tunneling Microscope,,,https://matvoc.nims.go.jp/entity/Q1949,2022-07-11 04:36:42.734137+00
f62e87bc-92aa-abc4-5471-cf467cb3912b,原子間力顕微鏡,Atomic Force Microscope,,,https://matvoc.nims.go.jp/entity/Q1950,2022-07-11 04:36:42.734137+00
e5cdc84b-5fab-f2b2-beb9-ee978d564d4d,クロマトグラフ,Chromatograph,,,https://matvoc.nims.go.jp/entity/Q1893,2022-07-11 04:36:42.734137+00
3022d0e8-8d0e-5af7-576e-cd5a85b5630b,ガスクロマトグラフ,Gas-phase Chromatograph,,,https://matvoc.nims.go.jp/entity/Q1951,2022-07-11 04:36:42.734137+00
98431a8f-f9de-fcd5-2c3e-e92a4989e2af,イオンクロマトグラフ,Ion Chromatograph,,,https://matvoc.nims.go.jp/entity/Q1952,2022-07-11 04:36:42.734137+00
4d5f1b0b-cb28-a0f7-267f-4beb36ce312f,液体クロマトグラフ,Lliquid-phase Chromatograph,,,https://matvoc.nims.go.jp/entity/Q1953,2022-07-11 04:36:42.734137+00
e98fbe1d-3346-9c2a-0df6-6e74164c8c71,ゲル浸透クロマトグラフ,Gel Permeation Chromatograph ,,,https://matvoc.nims.go.jp/entity/Q1954,2022-07-11 04:36:42.734137+00
12098051-b72a-87c5-1207-d92f960b53ce,分光,Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1894,2022-07-11 04:36:42.734137+00
01c2ecf1-15b5-54a1-6ec6-6fd5298b3e6d,赤外分光,Infrared Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1955,2022-07-11 04:36:42.734137+00
fd8bdfa7-ac0a-3183-7ffd-bdcd3db75414,紫外・可視分光,Ultraviolet Visible Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1956,2022-07-11 04:36:42.734137+00
b28a6b91-b052-cb6b-9912-404783e99b01,紫外可視近赤外分光,Ultraviolet Visible Near-Infrared Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1957,2022-07-11 04:36:42.734137+00
360aafa3-2a5f-307a-d62b-d393396894fb,近赤外分光光度計,Nnear‐Infrared Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1958,2022-07-11 04:36:42.734137+00
ca72ee47-924e-2346-fe01-ac8ff51437dc,蛍光分光,Flourescence Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1959,2022-07-11 04:36:42.734137+00
1def78ea-86cc-c954-ccf8-a06c7d5adff8,誘導結合プラズマ発光分光分析計,Inductively Coupled Plasma Atomic Emission Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1960,2022-07-11 04:36:42.734137+00
c0936a0f-678a-ef7d-3a64-eda1c3082d2a,X線蛍光分光分析,X-Ray Flourescence Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1966,2022-07-11 04:36:42.734137+00
3aad993f-ed09-bf50-78d6-b5a0f768fe74,ラマン分光,Raman Spectroscopy ,,,https://matvoc.nims.go.jp/entity/Q1962,2022-07-11 04:36:42.734137+00
7c46b215-c32e-ca99-3d2d-553375818dc8,円二色性分光,Circular Dichroism,,,https://matvoc.nims.go.jp/entity/Q1963,2022-07-11 04:36:42.734137+00
594ebf91-c422-6c96-69e0-67993820ebaf,X線吸収分光,X-Ray Absorption Spectroscopy ,,,https://matvoc.nims.go.jp/entity/Q2114,2022-07-11 04:36:42.734137+00
4408a09b-4afe-f469-8e59-11c5f6e8de21,X線発光分光,X-Ray Emission Spectroscopy ,,,https://matvoc.nims.go.jp/entity/Q1989,2022-07-11 04:36:42.734137+00
713666fe-b517-9271-e8ad-9ea9fd8d425a,X線光電子分光,X-Ray Photoelectron  Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1972,2022-07-11 04:36:42.734137+00
a89719fd-06d1-f3fc-355d-20c74fb5d41f,オージェ電子分光,Auger Electron  Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1967,2022-07-11 04:36:42.734137+00
7051bb5e-c223-b30d-e5d0-63fd7c9b77be,光電子分光,Photoemission Electron  Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q1968,2022-07-11 04:36:42.734137+00
8365d686-4134-cb63-8fad-7cbb49095241,走査型X線顕微鏡,Scanning X-Ray Microscope,,,https://matvoc.nims.go.jp/entity/Q1969,2022-07-11 04:36:42.734137+00
5f4213b0-bc8d-4e3e-72fb-972ff35c5033,放射光,Synchrotron Radiation,,,https://matvoc.nims.go.jp/entity/Q1896,2022-07-11 04:36:42.734137+00
e2ce4573-c1ca-69a2-df77-d65b0f552715,硬X線光電子分光法,HardX-ray Photoelectron Spectroscopy,,,https://matvoc.nims.go.jp/entity/Q2029,2022-07-11 04:36:42.734137+00
32be7300-40d6-b3c9-c88b-f07386c1c4de,装置・広域X線吸収微細構造,Extended X-ray Absorption Fine Structure,,,https://matvoc.nims.go.jp/entity/Q1971,2022-07-11 04:36:42.734137+00
88cabb25-3401-e1fe-f320-c517b04f63f8,X線吸収端近傍構造,X-ray Absorption Near Edge Structure,,,https://matvoc.nims.go.jp/entity/Q1964,2022-07-11 04:36:42.734137+00
03823a43-e4ab-8ae9-0a34-ad1758c39d0c,X線回折装置(放射光),X-ray Diffraction-Synchrotron Radiation,,,https://matvoc.nims.go.jp/entity/Q1961,2022-07-11 04:36:42.734137+00
ad23ce70-4d79-ef73-e810-f17d61be8d30,質量分析,Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1897,2022-07-11 04:36:42.734137+00
0ee6ea78-522c-18e7-3115-033dbb6cd885,二重収束質量分析　,Double-Focusing Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1974,2022-07-11 04:36:42.734137+00
d82dab4e-a2e4-1022-1513-3d1369c97d82,四重極質量分析　,Quadrupole Mass Analyzer; Quadrupole Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1975,2022-07-11 04:36:42.734137+00
bd783a01-03b8-a68c-e41c-2bbad3f782ce,飛行時間質量分析　,Time-Of-Flight Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1976,2022-07-11 04:36:42.734137+00
cb7d16df-4747-d79c-dd2f-2869cd14a62b,イオントラップ質量分析　,Ion Trap Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1977,2022-07-11 04:36:42.734137+00
4e1b9dfe-9d4e-c85b-b7d7-00c62e8e0f1f,フーリエ変換イオンサイクロトロン共鳴質量分析　,Fourier Transfom Ion Cyclotron Resonance Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1978,2022-07-11 04:36:42.734137+00
94b0ea79-4499-ef65-a4ee-142095f95527,飛行時間二次イオン質量分析　,Time-Of-Flight Secondary Ion Mass Spectrometry,,,https://matvoc.nims.go.jp/entity/Q1979,2022-07-11 04:36:42.734137+00
99d4cf2a-fd94-4aec-db0c-20905a9e0110,誘導結合プラズマ質量分析　,Inductively Coupled Plasma Mass Spectrometry,,,https://matvoc.nims.go.jp/entity/Q1980,2022-07-11 04:36:42.734137+00
de7d4b51-1294-aed2-2b00-52e9a6b68e2c,マトリックス支援レーザー脱離イオン化質量分析,Maldi-Tof Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1981,2022-07-11 04:36:42.734137+00
725d3eac-3ec5-6653-bb0c-9fe49a309d02,二次イオン質量分析,Secondary Ion Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q1982,2022-07-11 04:36:42.734137+00
3263abcd-9e15-3ba1-1d6a-5b13b954b3ac,圧縮試験,Compression Test,,,https://matvoc.nims.go.jp/entity/Q2025,2022-07-11 04:36:42.734137+00
926f3169-e302-ed92-770b-cbc2828e685e,直接イオン化質量分析,Direct Analysis In Real Time Mass Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2064,2022-07-11 04:36:42.734137+00
0de635ed-fe5b-63fa-2b02-aacb503ae4d5,ガスクロマトグラフ質量分析　,Gas Chromatography - Mass Spectorometer ,,,https://matvoc.nims.go.jp/entity/Q1984,2022-07-11 04:36:42.734137+00
9dff9f03-5814-233f-bde6-431b633f6d9d,液体クロマトグラフ質量分析　,Liquid Chromatography - Mass Spectorometer ,,,https://matvoc.nims.go.jp/entity/Q1985,2022-07-11 04:36:42.734137+00
6f44495b-841f-bcef-cc54-19684673500d,回折・散乱,Scattering & Diffraction ,,,https://matvoc.nims.go.jp/entity/Q1898,2022-07-11 04:36:42.734137+00
3b92a0c1-805c-f355-4bb3-b37b3005a124,X線回折装置,X-Ray Diffraction,,,https://matvoc.nims.go.jp/entity/Q1965,2022-07-11 04:36:42.734137+00
1c36274a-4a39-32e8-59cd-0dee6215b609,単結晶X線回折,Single Crystal X-Ray Diffraction,,,https://matvoc.nims.go.jp/entity/Q1986,2022-07-11 04:36:42.734137+00
07493a76-c527-4cb5-79ff-d716460cfc65,中性子回折,Neutron Diffraction,,,https://matvoc.nims.go.jp/entity/Q1987,2022-07-11 04:36:42.734137+00
8230ebf2-26fe-d94d-37af-8b4aa3e7ef2e,X線トポグラフィー,X-Ray Topography,,,https://matvoc.nims.go.jp/entity/Q1973,2022-07-11 04:36:42.734137+00
c2b58ea8-1409-5169-ae0c-921d7a6d95e9,X線マイクロトモグラフィー,X-Ray Microtomography,,,https://matvoc.nims.go.jp/entity/Q1988,2022-07-11 04:36:42.734137+00
d48017cc-5673-fef3-6f9e-4d6a4db3bd83,ラザフォード後方散乱,Rutherford Backscattering Spectrometry,,,https://matvoc.nims.go.jp/entity/Q1990,2022-07-11 04:36:42.734137+00
94f08db2-83ea-fb03-2f25-4e0136fbe1b3,電子回折,Electron Diffraction,,,https://matvoc.nims.go.jp/entity/Q1991,2022-07-11 04:36:42.734137+00
1dd1d570-026a-50f0-0246-b239fa67bd7d,磁気特性,Magnetic Characteristic,,,https://matvoc.nims.go.jp/entity/Q1899,2022-07-11 04:36:42.734137+00
e8df29a4-5ea2-7aff-0e28-3395aaba5977,磁気特性測定システム,Magnetic Property Measurement System,,,https://matvoc.nims.go.jp/entity/Q1992,2022-07-11 04:36:42.734137+00
c75d8bea-eb3f-89f0-5def-d17ce6e233dd,物理特性測定装置,Physical Property Measurement System,,,https://matvoc.nims.go.jp/entity/Q1993,2022-07-11 04:36:42.734137+00
10be6760-cb99-7ccd-d2ba-680e2342658b,振動試料型磁束計,Vibrating Sample Magnetometer,,,https://matvoc.nims.go.jp/entity/Q1994,2022-07-11 04:36:42.734137+00
0e18cddd-933d-75b5-df32-12b248b1849c,バイオ装置,Biological,,,https://matvoc.nims.go.jp/entity/Q1900,2022-07-11 04:36:42.734137+00
1bc28a78-67e8-49eb-6fbf-6a783f8b118a,リアルタイムPCR装置,Real-Time PCR,,,https://matvoc.nims.go.jp/entity/Q1995,2022-07-11 04:36:42.734137+00
f4501971-df1c-8747-a28f-af9461076603,PCR装置,PCR,,,https://matvoc.nims.go.jp/entity/Q1996,2022-07-11 04:36:42.734137+00
d7f4b22f-01f1-bc2f-c126-cfe8e12b0733,表面プラズモン共鳴装置,Surface Plasmon Resonance (SPR),,,https://matvoc.nims.go.jp/entity/Q1997,2022-07-11 04:36:42.734137+00
47029be4-87f5-86cc-d377-c2ca02f86b5a,プレートリーダー,Plate Reader,,,https://matvoc.nims.go.jp/entity/Q1998,2022-07-11 04:36:42.734137+00
7437fbfb-f560-a6fc-cffc-59c71e545df1,レーザースキャナー,Laser Scanner,,,https://matvoc.nims.go.jp/entity/Q1999,2022-07-11 04:36:42.734137+00
cc234f68-9550-8e9a-498f-d2facd27d695,フローサイトメトリー,Flow Cytometry,,,https://matvoc.nims.go.jp/entity/Q2000,2022-07-11 04:36:42.734137+00
8e62e286-74e9-4bb5-2f14-92da1307d85a,セルソーター,Cell Sorter,,,https://matvoc.nims.go.jp/entity/Q2001,2022-07-11 04:36:42.734137+00
3b7eb9da-a70e-a642-ffc0-58088896e031,電気泳動装置,Electrophoresis,,,https://matvoc.nims.go.jp/entity/Q2002,2022-07-11 04:36:42.734137+00
add609ce-b567-b8b0-6ad8-985f437ae67f,ゲルイメージング装置,Gel Imaging Device,,,https://matvoc.nims.go.jp/entity/Q2003,2022-07-11 04:36:42.734137+00
f23cb703-edf0-d6eb-45a7-d33559403c63,レーザーマイクロダイセクション,Laser Microdissection (LMD),,,https://matvoc.nims.go.jp/entity/Q2004,2022-07-11 04:36:42.734137+00
95b4e6c9-ff2d-0720-2a02-28943f3c40d3,DNAシーケンサー,DNA Sequencer,,,https://matvoc.nims.go.jp/entity/Q2069,2022-07-11 04:36:42.734137+00
8b7640a6-e0bf-6f06-b922-cd326b6efb85,その他分析装置,Analysis,,,https://matvoc.nims.go.jp/entity/Q1901,2022-07-11 04:36:42.734137+00
a2f98642-c00f-cd3d-1ef8-42c4d7ae7273,示差走査熱量分析,Differential Scanning Calorimetry,,,https://matvoc.nims.go.jp/entity/Q2005,2022-07-11 04:36:42.734137+00
c888a3e5-4289-910f-eb48-c608b62ae1d9,熱重量分析,Thermal Gravimetric Analysis,,,https://matvoc.nims.go.jp/entity/Q2007,2022-07-11 04:36:42.734137+00
1279c879-6248-f95c-31b7-9326af7c4437,示差熱・熱重量同時測定,Thermal Gravimetric Differential Scanning Calorimetry,,,https://matvoc.nims.go.jp/entity/Q2008,2022-07-11 04:36:42.734137+00
6bf2aae3-fc95-8f75-c3d1-02ad10eb6247,熱機械分析,Thermomechanical Analyzer,,,https://matvoc.nims.go.jp/entity/Q2009,2022-07-11 04:36:42.734137+00
598e6174-490e-071a-61bc-99c2dde91ee1,粘弾性測定,Viscoelasticity,,,https://matvoc.nims.go.jp/entity/Q2010,2022-07-11 04:36:42.734137+00
e60fa979-8ebc-939b-1598-5a1bdd417c94,段差計,Profiler,,,https://matvoc.nims.go.jp/entity/Q2011,2022-07-11 04:36:42.734137+00
9262aab9-c744-88c9-f733-e86808781737,膜厚測定,Film Thickness Measurement,,,https://matvoc.nims.go.jp/entity/Q1905,2022-07-11 04:36:42.734137+00
7989a34c-db48-246b-819a-858afa68c320,エリプソメーター,Ellipsometry,,,https://matvoc.nims.go.jp/entity/Q2013,2022-07-11 04:36:42.734137+00
2d3e316d-dd35-29ee-08ad-fd3823d7a8d5,接触角計,Contact Angle Meter ,,,https://matvoc.nims.go.jp/entity/Q2014,2022-07-11 04:36:42.734137+00
f13d554d-312e-84da-bb53-503a1b83c4a8,ゼータ電位,Zeta Potential,,,https://matvoc.nims.go.jp/entity/Q2015,2022-07-11 04:36:42.734137+00
8a0d08d1-fd78-40ef-863b-ba44fa677679,粒度分布測定（動的光散乱）,Dynamic Light Scattering ,,,https://matvoc.nims.go.jp/entity/Q2016,2022-07-11 04:36:42.734137+00
966072e7-7240-f3d7-daf4-828b7487823a,粒度分布測定（静的光散乱）,Static Light Scattering ,,,https://matvoc.nims.go.jp/entity/Q2017,2022-07-11 04:36:42.734137+00
e8b0dcc6-1f15-922e-2e8b-c79805365118,蒸気圧式絶対分子量測定,Vapor Pressure Osmometer,,,https://matvoc.nims.go.jp/entity/Q2018,2022-07-11 04:36:42.734137+00
6803eaa1-665b-7a38-8aa5-366a1ddbf4cb,電子物性評価,Electronic Property,,,https://matvoc.nims.go.jp/entity/Q2019,2022-07-11 04:36:42.734137+00
7d7175d7-854b-a755-11f0-2ee943529a60,電子材料・デバイス評価,Electronic Materials & Device characterization,,,https://matvoc.nims.go.jp/entity/Q2020,2022-07-11 04:36:42.734137+00
4f4f5685-b2ea-2e67-b2b1-7abf26f4e783,メスバウアー分光,Mössbauer Spectrometer,,,https://matvoc.nims.go.jp/entity/Q2021,2022-07-11 04:36:42.734137+00
b3dcd540-8cdf-b9c1-cbe1-2e8216602f44,電気化学,Electron Chemical,,,https://matvoc.nims.go.jp/entity/Q1902,2022-07-11 04:36:42.734137+00
adff1337-48fb-9f45-37f0-de44ae226633,電流滴定,Amoperometry,,,https://matvoc.nims.go.jp/entity/Q2022,2022-07-11 04:36:42.734137+00
c799b331-1491-54af-46e8-9aa9e85a7c4b,電位差測定,Potentiometry,,,https://matvoc.nims.go.jp/entity/Q2023,2022-07-11 04:36:42.734137+00
ed565c88-79d2-4f8f-e8a0-0adf8e6a6534,電流測定,Voltammetry,,,https://matvoc.nims.go.jp/entity/Q2024,2022-07-11 04:36:42.734137+00
bdc34fc6-b18c-3875-f5eb-8ce3c8840057,機械特性,Mechanical Properties,,,https://matvoc.nims.go.jp/entity/Q1903,2022-07-11 04:36:42.734137+00
59172962-ea0e-95af-792a-cf4ae7bedcbd,クリープ試験,Creep Test,,,https://matvoc.nims.go.jp/entity/Q2026,2022-07-11 04:36:42.734137+00
c03fdeea-75f7-fef8-2440-7295a75e8eca,動的機械分析,Dynamic Mechanical Analysis,,,https://matvoc.nims.go.jp/entity/Q2027,2022-07-11 04:36:42.734137+00
7f646c8d-de85-fc2c-4a2d-9a1bc38fd401,疲労試験,Fatigue Testing,,,https://matvoc.nims.go.jp/entity/Q2028,2022-07-11 04:36:42.734137+00
e40de2c3-b59f-3cf2-a849-0a5f438691bd,硬度計,Hardness Testing,,,https://matvoc.nims.go.jp/entity/Q1970,2022-07-11 04:36:42.734137+00
e3494e3d-0535-91f2-5293-7c0cfb4f31ec,ナノインデンテーション試験,Nanoindentation,,,https://matvoc.nims.go.jp/entity/Q2030,2022-07-11 04:36:42.734137+00
c043a655-36c9-311a-41b1-afa9757d6107,せん断　ねじれ,Shear or Torsion ,,,https://matvoc.nims.go.jp/entity/Q2031,2022-07-11 04:36:42.734137+00
630ee7e1-97b6-2315-3f06-af7515431897,引っ張り試験,Tension Test,,,https://matvoc.nims.go.jp/entity/Q2032,2022-07-11 04:36:42.734137+00
4bbc1d52-da05-f095-8f1d-31bd0a69b46f,計算,Calculation,,,https://matvoc.nims.go.jp/entity/Q1913,2022-07-11 04:36:42.734137+00
964f9f3f-bf68-28e2-5cb2-0c67689fbdd8,理論計算・シミュレーション,"Theory Calculation,Simulation",,,https://matvoc.nims.go.jp/entity/Q1904,2022-07-11 04:36:42.734137+00
486feb96-191a-74b5-76d4-52c641f24c0a,理論計算, Theoritical Calculation,,,https://matvoc.nims.go.jp/entity/Q2109,2022-07-11 04:36:42.734137+00
ebfbb965-7b4b-df13-0263-7b4e47a28ffc,シミュレーション,Simulation,,,https://matvoc.nims.go.jp/entity/Q2058,2022-07-11 04:36:42.734137+00
241a5166-d0f4-7834-495e-b4f44852852a,CAD,Computer-Aided Design,,,https://matvoc.nims.go.jp/entity/Q2033,2022-07-11 04:36:42.734137+00
1c58e25b-976d-884e-57c0-df7777e1b911,機械学習,Machine Learning,,,https://matvoc.nims.go.jp/entity/Q2034,2022-07-11 04:36:42.734137+00
16a0dea1-6815-a98d-12ab-4c8b264e1492,合成・プロセス装置,Synthesis and Processing Instruments,,,https://matvoc.nims.go.jp/entity/Q1914,2022-07-11 04:36:42.734137+00
a6bee1ed-79c5-0e51-5baa-aab2c3090183,蒸着・成膜装置,"Film formation, Deposition",,,https://matvoc.nims.go.jp/entity/Q2012,2022-07-11 04:36:42.734137+00
aa874ed9-ad8d-21ef-552f-8a1942675bf1,原子層堆積(ALD)装置,Atomic Layer Deposition System,,,https://matvoc.nims.go.jp/entity/Q2035,2022-07-11 04:36:42.734137+00
3b26b536-fd4a-6df7-0a0e-029246dcc16b,コーター,Coater,,,https://matvoc.nims.go.jp/entity/Q2036,2022-07-11 04:36:42.734137+00
17ea69fc-df0c-47a3-7c0d-aaf4646664d3,化学蒸着(CVD)装置,Chemical Vapor Deposition System,,,https://matvoc.nims.go.jp/entity/Q2037,2022-07-11 04:36:42.734137+00
6fdbb537-38b2-905e-e34f-77f7525b0a2a,電着装置,Electrodeposition System,,,https://matvoc.nims.go.jp/entity/Q2038,2022-07-11 04:36:42.734137+00
261b2899-48cd-f0d8-689f-b68edeb26d66,物理蒸着(PVD)装置,Physical Vapor Deposition System,,,https://matvoc.nims.go.jp/entity/Q2039,2022-07-11 04:36:42.734137+00
1703dc22-1da0-c5b7-92aa-a127acb2b588,インクジェット堆積装置,Ink-Jet Deposition System,,,https://matvoc.nims.go.jp/entity/Q2040,2022-07-11 04:36:42.734137+00
9da4fb00-3ca8-e68f-ad9e-3724af25bfe4,ラングミュア - ブロジェット膜堆積装置,Langmuir-Blodgett Film Deposition System,,,https://matvoc.nims.go.jp/entity/Q2041,2022-07-11 04:36:42.734137+00
745c14bb-d34a-8dfa-bd0e-d9a84fb56460,プラズマ溶射装置,Plasma Spray System,,,https://matvoc.nims.go.jp/entity/Q2042,2022-07-11 04:36:42.734137+00
bdddc620-e6dc-0df9-a4ee-30277f1bcc93,スッパタリング（スパッタ）,Sputtering,,,https://matvoc.nims.go.jp/entity/Q2043,2022-07-11 04:36:42.734137+00
9dc8b43b-18f0-44b6-c5d2-76fb844a174e,成形装置,"Molding,Forming",,,https://matvoc.nims.go.jp/entity/Q1916,2022-07-11 04:36:42.734137+00
1a0effaf-3afa-ad1a-2b77-15c7f774e26a,冷間圧延ローラー,Cold Rollers,,,https://matvoc.nims.go.jp/entity/Q2044,2022-07-11 04:36:42.734137+00
98f5224f-05a6-493b-d187-4b05be541af5,引抜金型,Drawing Die,,,https://matvoc.nims.go.jp/entity/Q2045,2022-07-11 04:36:42.734137+00
6a8a60ff-b96b-1819-3de7-ce60cfa024d0,押出金型,Extrusion Die,,,https://matvoc.nims.go.jp/entity/Q2046,2022-07-11 04:36:42.734137+00
ff9fa4eb-3252-9490-1527-82d461ed2718,鍛造機械,Forging Equipment,,,https://matvoc.nims.go.jp/entity/Q2047,2022-07-11 04:36:42.734137+00
bb82de8a-2302-2661-6ccf-d7837d7057bc,ホットプレス,Hot Press,,,https://matvoc.nims.go.jp/entity/Q2048,2022-07-11 04:36:42.734137+00
032a6ca0-c366-85e0-7e99-6adc47e3c7a3,熱間圧延ローラー,Hot Rolling,,,https://matvoc.nims.go.jp/entity/Q2049,2022-07-11 04:36:42.734137+00
0143e014-fad3-2e8c-801a-926e408a0a05,粉砕機,Mill,,,https://matvoc.nims.go.jp/entity/Q2050,2022-07-11 04:36:42.734137+00
49cb3a34-99c5-6d2b-7a40-e2a5e1f7401d,鋳型,Molding,,,https://matvoc.nims.go.jp/entity/Q2110,2022-07-11 04:36:42.734137+00
b240ce65-8f2b-93cd-adcc-9d9efac455a6,3Dプリンタ,3D Printer,,,https://matvoc.nims.go.jp/entity/Q2051,2022-07-11 04:36:42.734137+00
e83fea69-d559-dbf9-3bac-54baba3c4ef0,リソグラフィ,Lithography,,,https://matvoc.nims.go.jp/entity/Q1917,2022-07-11 04:36:42.734137+00
db29daf1-9110-36fb-abc8-3800bedb14be,光露光（マスクアライナ）,Mask Aligner,,,https://matvoc.nims.go.jp/entity/Q2052,2022-07-11 04:36:42.734137+00
61e1c4a0-c646-ffad-276b-34491b7b44dd,光露光（ステッパ）,Stepper,,,https://matvoc.nims.go.jp/entity/Q2053,2022-07-11 04:36:42.734137+00
f20b6729-0916-671d-37d5-2c996184d526,光露光（マスクレス、直接描画）,Maskless Exposure System,,,https://matvoc.nims.go.jp/entity/Q2054,2022-07-11 04:36:42.734137+00
e826ffd0-412c-94f6-fdf5-bed7e00fe9af,電子線描画（EB）,Electron Beam Lithography,,,https://matvoc.nims.go.jp/entity/Q2055,2022-07-11 04:36:42.734137+00
ba60962a-7513-b3a2-35fe-47c97df06535,ナノインプリント,Nanoimprint Lithography,,,https://matvoc.nims.go.jp/entity/Q2056,2022-07-11 04:36:42.734137+00
cdd484ac-8e28-4fd7-6f64-6e8ed02740f0,膜加工・エッチング,Etching,,,https://matvoc.nims.go.jp/entity/Q1918,2022-07-11 04:36:42.734137+00
4136d70b-55ea-66b6-86f3-649c42c3196c,ドライエッチング（RIE）,Dry Etching(Reactive Ion Etching),,,https://matvoc.nims.go.jp/entity/Q2057,2022-07-11 04:36:42.734137+00
49de9448-5281-2018-684c-4f665bb19953,ドライエッチング（ECR）,Dry Etching(Electron Cyclotron Resonance-RIE),,,https://matvoc.nims.go.jp/entity/Q2059,2022-07-11 04:36:42.734137+00
8fe177e6-cbad-d31d-304c-c61a459a1703,ドライエッチング（その他）,Dry Etching(Others),,,https://matvoc.nims.go.jp/entity/Q2060,2022-07-11 04:36:42.734137+00
e8cddaaf-3b36-1fb4-0749-f389088e40fc,ウェット／ガスエッチング,Wet Etching/Gas Etching,,,https://matvoc.nims.go.jp/entity/Q2061,2022-07-11 04:36:42.734137+00
ff720d56-381a-7b0f-a308-3912b3630caa,レーザー加工,Laser Processing,,,https://matvoc.nims.go.jp/entity/Q2062,2022-07-11 04:36:42.734137+00
f15fb97f-ba7f-8733-3862-b847511b1728,その他加工装置,Processing,,,https://matvoc.nims.go.jp/entity/Q1919,2022-07-11 04:36:42.734137+00
916e4a01-1b70-433f-344d-ae06716c9677,酸化,Oxidization System,,,https://matvoc.nims.go.jp/entity/Q2063,2022-07-11 04:36:42.734137+00
d38611ff-baee-02b8-c3e3-2984cbb452ec,拡散,Diffusion System,,,https://matvoc.nims.go.jp/entity/Q2006,2022-07-11 04:36:42.734137+00
a6dc608d-330d-2399-0d56-635163221e2e,イオン注入,Ion Implantation,,,https://matvoc.nims.go.jp/entity/Q2065,2022-07-11 04:36:42.734137+00
1720c881-c98c-1d0c-b2c7-4822326679bd,接合,Bonder,,,https://matvoc.nims.go.jp/entity/Q2066,2022-07-11 04:36:42.734137+00
7951f619-5c3a-6d47-ea9f-721d448585c2,レジスト塗布,Photoresist Spin Coater,,,https://matvoc.nims.go.jp/entity/Q2067,2022-07-11 04:36:42.734137+00
f8b19a20-1f10-e948-12c6-093e3145b0c1,現像装置,Photoresist Developer,,,https://matvoc.nims.go.jp/entity/Q2068,2022-07-11 04:36:42.734137+00
f20292de-998e-e192-da8b-95e4ab800ed8,合成設備,Synthesis,,,https://matvoc.nims.go.jp/entity/Q1920,2022-07-11 04:36:42.734137+00
49591a89-a59d-8d3f-d7cd-a4742f239339,分注機,Dispenser,,,https://matvoc.nims.go.jp/entity/Q1983,2022-07-11 04:36:42.734137+00
ebc1ce7d-6359-3d35-0e0b-73e848504393,遠心機,Centrifuge,,,https://matvoc.nims.go.jp/entity/Q2070,2022-07-11 04:36:42.734137+00
4a684b87-c1db-005f-9935-f75b4e2f9056,撹拌機,Stirrer,,,https://matvoc.nims.go.jp/entity/Q2071,2022-07-11 04:36:42.734137+00
33c6e9dc-5787-0f96-7683-f39281c60419,化学式、組成式、分子式など,"Chemical formula, composition formula, molecular formula, etc.",化学式、組成式、分子式などを入力してください,"Please enter Chemical formula, composition formula, molecular formula, etc.",NULL,2022-07-11 04:36:42.734137+00
f2d5e89e-01f0-66a2-5d8e-623a4fc31698,物質名,Material name,物質名を入力してください,Please enter Material name,NULL,2022-07-11 04:36:42.734137+00
a7a6fc7b-ed46-88b0-bba8-a1e34857a049,試料別名,Another sample name,試料別名を入力してください,Please enter Another sample name,NULL,2022-07-11 04:36:42.734137+00
f207e704-9308-42f0-b090-98e2db81c757,グロー放電質量分析法,Glow Discharge Mass Spectrometry,NULL,NULL,https://matvoc.nims.go.jp/entity/Q2839,2022-09-05 04:02:31.114117+00
0d0417a3-3c3b-496a-b0fb-5a26f8a74166,ロット番号、製造番号など,Lot number or product number etc,,,NULL,2022-10-11 06:13:06.860778+00
e2d20d02-2e38-2cd3-b1b3-66fdb8a11057,CAS番号,CAS Number,CAS番号を入力してください,Please enter CAS Number,NULL,2022-07-11 04:36:42.734137+00
1e70d11d-cbdd-bfd1-9301-9612c29b4060,試料購入日,Purchase date,試料購入日を入力してください,Please enter Purchase date,NULL,2022-07-11 04:36:42.734137+00
1d3cab05-3eaa-cb9b-9a3f-20eb0ca26963,結晶状態,Crystalline state,結晶状態を入力してください,Please enter Crystalline state,NULL,2022-07-11 04:36:42.734137+00
efcf34e7-4308-c195-6691-6f4d28ffc9bb,結晶構造,Crystal structure,結晶構造を入力してください,Please enter Crystal structure,NULL,2022-07-11 04:36:42.734137+00
e9617207-7f74-ef45-9b05-74eef6e4ecbb,ピアソン記号,Pearson symbol,ピアソン記号を入力してください,Please enter Pearson symbol,NULL,2022-07-11 04:36:42.734137+00
f63149a4-e57c-4273-4c1e-dffa41356d28,空間群,Space group,空間群を入力してください,Please enter Space group,https://matvoc.nims.go.jp/wiki/Item:Q224,2022-07-11 04:36:42.734137+00
7cc57dfb-8b70-4b3a-5315-fbce4cbf73d0,試料形状,Sample shape,試料形状を入力してください,Please enter Sample shape,NULL,2022-07-11 04:36:42.734137+00
3250c45d-0ed6-1438-43b5-eb679918604a,化学式,Chemical formula,化学式を入力してください,Please enter Chemical formula,NULL,2022-07-11 04:36:42.734137+00
70c2c751-5404-19b7-4a5e-981e6cebbb15,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00
518e26a0-4262-86f5-3598-80e18e6ff2af,PubChem,PubChem,PubChemを入力してください,Please enter PubChem,NULL,2022-07-11 04:36:42.734137+00
3a775d54-5c13-fe66-6405-29c05bc931ce,粘度,viscosity,粘度を入力してください,Please enter viscosity,https://matvoc.nims.go.jp/wiki/Item:Q284,2022-07-11 04:36:42.734137+00
659da80e-c2ee-2986-41ce-68201b3bc4dd,沸点,boiling point,沸点を入力してください,Please enter boiling point,NULL,2022-07-11 04:36:42.734137+00
4efc4c3b-727c-c752-cf28-701b55dba1af,融点,Melting temperature,融点を入力してください,Please enter Melting temperature,https://matvoc.nims.go.jp/wiki/Item:Q297,2022-07-11 04:36:42.734137+00
dc27a956-263e-f920-e574-5beec912a247,分子量,molecular weight,分子量を入力してください,Please enter molecular weight,https://matvoc.nims.go.jp/entity/Q551,2022-08-01 09:29:30.179916+00
efc6a0d5-313e-1871-190c-baaff7d1bf6c,SMILES String,SMILES String,SMILES Stringを入力してください,Please enter SMILES String,NULL,2022-08-01 09:32:47.462801+00
3edadcff-8a85-51d9-708f-8f76bf055377,InChI key,InChI key,InChI keyを入力してください,Please enter InChI key,NULL,2022-08-01 09:32:47.462801+00
0444cf53-db47-b208-7b5f-54429291a140,試料分類,Sample type,試料分類を入力してください,Please enter Sample type,NULL,2022-08-09 02:57:35.896387+00
fc30c31d-12a3-591a-c837-4f06ab458de0,生物種,Taxonomy,生物種を入力してください,Please enter Taxonomy,NULL,2022-08-09 02:57:35.896387+00
9a23002a-c398-e521-081a-24b6cd32dbbd,細胞株,Cell line,細胞株を入力してください,Please enter Cell line,NULL,2022-08-09 02:57:35.896387+00
b4ce4016-e2bf-e5a1-7cae-ed496c7a776f,タンパク名,Protein name,タンパク名を入力してください,Please enter Protein name,NULL,2022-08-09 02:57:35.896387+00
8c9b1a88-1530-24d3-4b2e-5441eee5c24f,遺伝子名,Gene name,遺伝子名を入力してください,Please enter Gene name,NULL,2022-08-09 02:57:35.896387+00
047e30f3-f294-e58d-cbe4-6bb588bf4cf8,NCBIアクセッション番号,NCBI accession number,NCBIアクセッション番号を入力してください,Please enter NCBI accession number,NULL,2022-08-09 02:57:35.896387+00
3adf9874-7bcb-e5f8-99cb-3d6fd9d7b55e,一般名称,General name,一般名称を入力してください,Please enter General name,NULL,2022-08-24 01:04:54.879136+00
9270879d-d94e-4d3f-2d5c-19568e040004,InChI,InChI,InChIを入力してください,Please enter InChI,NULL,2022-08-01 09:32:47.462801+00
    """.strip()


def csv_sample_general_sample_term() -> str:
    """csv_sample_general_sample_termのCSVデータを返す機能

    TODO
     - 本来は、term_id,key_nameだけでよく、それ以降は別のテーブルを参照しているもの。ここも2列のデータでよい。
    """

    return """
term_id,key_name,dict.term.name_ja,dict.term.name_en,dict.term.hint_ja,dict.term.hint_en,dict.term.term_uri,dict.term.created
33c6e9dc-5787-0f96-7683-f39281c60419,sample.general.composiiton,化学式、組成式、分子式など,"Chemical formula, composition formula, molecular formula, etc.",化学式、組成式、分子式などを入力してください,"Please enter Chemical formula, composition formula, molecular formula, etc.",NULL,2022-07-11 04:36:42.734137+00
f2d5e89e-01f0-66a2-5d8e-623a4fc31698,sample.general.material-name,物質名,Material name,物質名を入力してください,Please enter Material name,NULL,2022-07-11 04:36:42.734137+00
a7a6fc7b-ed46-88b0-bba8-a1e34857a049,sample.general.sample-alias,試料別名,Another sample name,試料別名を入力してください,Please enter Another sample name,NULL,2022-07-11 04:36:42.734137+00
e2d20d02-2e38-2cd3-b1b3-66fdb8a11057,sample.general.cas-number,CAS番号,CAS Number,CAS番号を入力してください,Please enter CAS Number,NULL,2022-07-11 04:36:42.734137+00
1e70d11d-cbdd-bfd1-9301-9612c29b4060,sample.general.purchase-date,試料購入日,Purchase date,試料購入日を入力してください,Please enter Purchase date,NULL,2022-07-11 04:36:42.734137+00
1d3cab05-3eaa-cb9b-9a3f-20eb0ca26963,sample.general.crystalline-state,結晶状態,Crystalline state,結晶状態を入力してください,Please enter Crystalline state,NULL,2022-07-11 04:36:42.734137+00
efcf34e7-4308-c195-6691-6f4d28ffc9bb,sample.general.crystal-structure,結晶構造,Crystal structure,結晶構造を入力してください,Please enter Crystal structure,NULL,2022-07-11 04:36:42.734137+00
e9617207-7f74-ef45-9b05-74eef6e4ecbb,sample.general.pearson-symbol,ピアソン記号,Pearson symbol,ピアソン記号を入力してください,Please enter Pearson symbol,NULL,2022-07-11 04:36:42.734137+00
f63149a4-e57c-4273-4c1e-dffa41356d28,sample.general.space-group,空間群,Space group,空間群を入力してください,Please enter Space group,https://matvoc.nims.go.jp/wiki/Item:Q224,2022-07-11 04:36:42.734137+00
7cc57dfb-8b70-4b3a-5315-fbce4cbf73d0,sample.general.sample-shape,試料形状,Sample shape,試料形状を入力してください,Please enter Sample shape,NULL,2022-07-11 04:36:42.734137+00
efc6a0d5-313e-1871-190c-baaff7d1bf6c,sample.general.smiles-string,SMILES String,SMILES String,SMILES Stringを入力してください,Please enter SMILES String,NULL,2022-08-01 09:32:47.462801+00
9270879d-d94e-4d3f-2d5c-19568e040004,sample.general.inchi,InChI,InChI,InChIを入力してください,Please enter InChI,NULL,2022-08-01 09:32:47.462801+00
3edadcff-8a85-51d9-708f-8f76bf055377,sample.general.inchi-key,InChI key,InChI key,InChI keyを入力してください,Please enter InChI key,NULL,2022-08-01 09:32:47.462801+00
dc27a956-263e-f920-e574-5beec912a247,sample.general.molecular-weight,分子量,molecular weight,分子量を入力してください,Please enter molecular weight,https://matvoc.nims.go.jp/entity/Q551,2022-08-01 09:29:30.179916+00
0444cf53-db47-b208-7b5f-54429291a140,sample.general.sample-type,試料分類,Sample type,試料分類を入力してください,Please enter Sample type,NULL,2022-08-09 02:57:35.896387+00
fc30c31d-12a3-591a-c837-4f06ab458de0,sample.general.taxonomy,生物種,Taxonomy,生物種を入力してください,Please enter Taxonomy,NULL,2022-08-09 02:57:35.896387+00
9a23002a-c398-e521-081a-24b6cd32dbbd,sample.general.cell-line,細胞株,Cell line,細胞株を入力してください,Please enter Cell line,NULL,2022-08-09 02:57:35.896387+00
b4ce4016-e2bf-e5a1-7cae-ed496c7a776f,sample.general.protein-name,タンパク名,Protein name,タンパク名を入力してください,Please enter Protein name,NULL,2022-08-09 02:57:35.896387+00
8c9b1a88-1530-24d3-4b2e-5441eee5c24f,sample.general.gene-name,遺伝子名,Gene name,遺伝子名を入力してください,Please enter Gene name,NULL,2022-08-09 02:57:35.896387+00
047e30f3-f294-e58d-cbe4-6bb588bf4cf8,sample.general.ncbi-accession-number,NCBIアクセッション番号,NCBI accession number,NCBIアクセッション番号を入力してください,Please enter NCBI accession number,NULL,2022-08-09 02:57:35.896387+00
3adf9874-7bcb-e5f8-99cb-3d6fd9d7b55e,sample.general.general-name,一般名称,General name,一般名称を入力してください,Please enter General name,NULL,2022-08-24 01:04:54.879136+00
0aadfff2-37de-411f-883a-38b62b2abbce,sample.general.chemical-composition,化学組成,Chemical composition,,,NULL,2022-10-11 06:12:34.454294+00
5e166ac4-bfcd-457a-84bc-8626abe9188f,sample.general.supplier,購入元,Supplier,,,NULL,2022-10-11 06:12:52.876335+00
0d0417a3-3c3b-496a-b0fb-5a26f8a74166,sample.general.lot-number-or-product-number-etc,ロット番号、製造番号など,Lot number or product number etc,,,NULL,2022-10-11 06:13:06.860778+00
    """.strip()


def csv_sample_sample_class() -> str:
    """csv_sample_sample_classのデータを返す機能"""
    return """
id,name_ja,name_en
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,有機材料,organic material
932e4fe1-9724-305f-ffc5-1908c31c83e5,無機材料,inorganic material
a674a8ef-efa8-9497-4ed4-74de55fafddb,金属・合金,metals and alloys
342ba516-4d02-171c-9bc4-70a3134b47a8,ポリマー,polymers
52148afb-6759-23e8-c8b8-33912ec5bfcf,半導体,semiconductors
961c9637-9b83-0e9d-e60e-ffc1e2517afd,セラミックス,ceramics
0dde5969-3039-739b-b33b-97df40450790,生物学的物質,biological
    """.strip()


def csv_sample_specific_sample_term() -> str:
    """csv_sample_specific_sample_termのCSVデータを返す機能

    TODO
     - 本来は、sample_class_id,term_id,key_nameだけでよく、それ以降は別のテーブルを参照しているもの。ここも3列のデータでよい。
    """

    return """
sample_class_id,term_id,key_name,sample.sample_class.name_ja,sample.sample_class.name_en,dict.term.name_ja,dict.term.name_en,dict.term.hint_ja,dict.term.hint_en,dict.term.term_uri,dict.term.created,bind_class_and_term_ja,bind_class_and_term_en
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,3250c45d-0ed6-1438-43b5-eb679918604a,sample.specific.organic.chemical-formula,有機材料,organic material,化学式,Chemical formula,化学式を入力してください,Please enter Chemical formula,NULL,2022-07-11 04:36:42.734137+00,有機材料/化学式,organic material/Chemical formula
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.organic.name,有機材料,organic material,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,有機材料/名称,organic material/Name
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,e2d20d02-2e38-2cd3-b1b3-66fdb8a11057,sample.specific.organic.cas-number,有機材料,organic material,CAS番号,CAS Number,CAS番号を入力してください,Please enter CAS Number,NULL,2022-07-11 04:36:42.734137+00,有機材料/CAS番号,organic material/CAS Number
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,518e26a0-4262-86f5-3598-80e18e6ff2af,sample.specific.organic.pubchem,有機材料,organic material,PubChem,PubChem,PubChemを入力してください,Please enter PubChem,NULL,2022-07-11 04:36:42.734137+00,有機材料/PubChem,organic material/PubChem
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,3a775d54-5c13-fe66-6405-29c05bc931ce,sample.specific.organic.viscosity,有機材料,organic material,粘度,viscosity,粘度を入力してください,Please enter viscosity,https://matvoc.nims.go.jp/wiki/Item:Q284,2022-07-11 04:36:42.734137+00,有機材料/粘度,organic material/viscosity
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,659da80e-c2ee-2986-41ce-68201b3bc4dd,sample.specific.organic.boiling-point,有機材料,organic material,沸点,boiling point,沸点を入力してください,Please enter boiling point,NULL,2022-07-11 04:36:42.734137+00,有機材料/沸点,organic material/boiling point
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,4efc4c3b-727c-c752-cf28-701b55dba1af,sample.specific.organic.melting-temperature,有機材料,organic material,融点,Melting temperature,融点を入力してください,Please enter Melting temperature,https://matvoc.nims.go.jp/wiki/Item:Q297,2022-07-11 04:36:42.734137+00,有機材料/融点,organic material/Melting temperature
932e4fe1-9724-305f-ffc5-1908c31c83e5,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.inorganic.name,無機材料,inorganic material,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,無機材料/名称,inorganic material/Name
932e4fe1-9724-305f-ffc5-1908c31c83e5,3250c45d-0ed6-1438-43b5-eb679918604a,sample.specific.inorganic.chemical-formula,無機材料,inorganic material,化学式,Chemical formula,化学式を入力してください,Please enter Chemical formula,NULL,2022-07-11 04:36:42.734137+00,無機材料/化学式,inorganic material/Chemical formula
932e4fe1-9724-305f-ffc5-1908c31c83e5,f63149a4-e57c-4273-4c1e-dffa41356d28,sample.specific.inorganic.space-group,無機材料,inorganic material,空間群,Space group,空間群を入力してください,Please enter Space group,https://matvoc.nims.go.jp/wiki/Item:Q224,2022-07-11 04:36:42.734137+00,無機材料/空間群,inorganic material/Space group
a674a8ef-efa8-9497-4ed4-74de55fafddb,3250c45d-0ed6-1438-43b5-eb679918604a,sample.specific.metals.chemical-formula,金属・合金,metals and alloys,化学式,Chemical formula,化学式を入力してください,Please enter Chemical formula,NULL,2022-07-11 04:36:42.734137+00,金属・合金/化学式,metals and alloys/Chemical formula
a674a8ef-efa8-9497-4ed4-74de55fafddb,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.metals.name,金属・合金,metals and alloys,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,金属・合金/名称,metals and alloys/Name
a674a8ef-efa8-9497-4ed4-74de55fafddb,e2d20d02-2e38-2cd3-b1b3-66fdb8a11057,sample.specific.metals.cas-number,金属・合金,metals and alloys,CAS番号,CAS Number,CAS番号を入力してください,Please enter CAS Number,NULL,2022-07-11 04:36:42.734137+00,金属・合金/CAS番号,metals and alloys/CAS Number
a674a8ef-efa8-9497-4ed4-74de55fafddb,f63149a4-e57c-4273-4c1e-dffa41356d28,sample.specific.metals.space-group,金属・合金,metals and alloys,空間群,Space group,空間群を入力してください,Please enter Space group,https://matvoc.nims.go.jp/wiki/Item:Q224,2022-07-11 04:36:42.734137+00,金属・合金/空間群,metals and alloys/Space group
a674a8ef-efa8-9497-4ed4-74de55fafddb,efcf34e7-4308-c195-6691-6f4d28ffc9bb,sample.specific.metals.crystal-structure,金属・合金,metals and alloys,結晶構造,Crystal structure,結晶構造を入力してください,Please enter Crystal structure,NULL,2022-07-11 04:36:42.734137+00,金属・合金/結晶構造,metals and alloys/Crystal structure
a674a8ef-efa8-9497-4ed4-74de55fafddb,659da80e-c2ee-2986-41ce-68201b3bc4dd,sample.specific.metals.boiling-point,金属・合金,metals and alloys,沸点,boiling point,沸点を入力してください,Please enter boiling point,NULL,2022-07-11 04:36:42.734137+00,金属・合金/沸点,metals and alloys/boiling point
a674a8ef-efa8-9497-4ed4-74de55fafddb,4efc4c3b-727c-c752-cf28-701b55dba1af,sample.specific.metals.melting-temperature,金属・合金,metals and alloys,融点,Melting temperature,融点を入力してください,Please enter Melting temperature,https://matvoc.nims.go.jp/wiki/Item:Q297,2022-07-11 04:36:42.734137+00,金属・合金/融点,metals and alloys/Melting temperature
342ba516-4d02-171c-9bc4-70a3134b47a8,3250c45d-0ed6-1438-43b5-eb679918604a,sample.specific.polymers.chemical-formula,ポリマー,polymers,化学式,Chemical formula,化学式を入力してください,Please enter Chemical formula,NULL,2022-07-11 04:36:42.734137+00,ポリマー/化学式,polymers/Chemical formula
342ba516-4d02-171c-9bc4-70a3134b47a8,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.polymers.name,ポリマー,polymers,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,ポリマー/名称,polymers/Name
342ba516-4d02-171c-9bc4-70a3134b47a8,e2d20d02-2e38-2cd3-b1b3-66fdb8a11057,sample.specific.polymers.cas-number,ポリマー,polymers,CAS番号,CAS Number,CAS番号を入力してください,Please enter CAS Number,NULL,2022-07-11 04:36:42.734137+00,ポリマー/CAS番号,polymers/CAS Number
342ba516-4d02-171c-9bc4-70a3134b47a8,518e26a0-4262-86f5-3598-80e18e6ff2af,sample.specific.polymers.pubchem,ポリマー,polymers,PubChem,PubChem,PubChemを入力してください,Please enter PubChem,NULL,2022-07-11 04:36:42.734137+00,ポリマー/PubChem,polymers/PubChem
342ba516-4d02-171c-9bc4-70a3134b47a8,4efc4c3b-727c-c752-cf28-701b55dba1af,sample.specific.polymers.melting-temperature,ポリマー,polymers,融点,Melting temperature,融点を入力してください,Please enter Melting temperature,https://matvoc.nims.go.jp/wiki/Item:Q297,2022-07-11 04:36:42.734137+00,ポリマー/融点,polymers/Melting temperature
52148afb-6759-23e8-c8b8-33912ec5bfcf,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.semiconductors.name,半導体,semiconductors,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,半導体/名称,semiconductors/Name
961c9637-9b83-0e9d-e60e-ffc1e2517afd,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.ceramics.name,セラミックス,ceramics,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,セラミックス/名称,ceramics/Name
0dde5969-3039-739b-b33b-97df40450790,70c2c751-5404-19b7-4a5e-981e6cebbb15,sample.specific.biological.name,生物学的物質,biological,名称,Name,名称を入力してください,Please enter Name,NULL,2022-07-11 04:36:42.734137+00,生物学的物質/名称,biological/Name
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,dc27a956-263e-f920-e574-5beec912a247,sample.specific.organic-material.molecular-weight,有機材料,organic material,分子量,molecular weight,分子量を入力してください,Please enter molecular weight,https://matvoc.nims.go.jp/entity/Q551,2022-08-01 09:29:30.179916+00,有機材料/分子量,organic material/molecular weight
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,efc6a0d5-313e-1871-190c-baaff7d1bf6c,sample.specific.organic-material.SMILES-String,有機材料,organic material,SMILES String,SMILES String,SMILES Stringを入力してください,Please enter SMILES String,NULL,2022-08-01 09:32:47.462801+00,有機材料/SMILES String,organic material/SMILES String
0dde5969-3039-739b-b33b-97df40450790,0444cf53-db47-b208-7b5f-54429291a140,sample.specific.biological.sample-type,生物学的物質,biological,試料分類,Sample type,試料分類を入力してください,Please enter Sample type,NULL,2022-08-09 02:57:35.896387+00,生物学的物質/試料分類,biological/Sample type
0dde5969-3039-739b-b33b-97df40450790,fc30c31d-12a3-591a-c837-4f06ab458de0,sample.specific.biological.taxonomy,生物学的物質,biological,生物種,Taxonomy,生物種を入力してください,Please enter Taxonomy,NULL,2022-08-09 02:57:35.896387+00,生物学的物質/生物種,biological/Taxonomy
0dde5969-3039-739b-b33b-97df40450790,9a23002a-c398-e521-081a-24b6cd32dbbd,sample.specific.biological.cell-line,生物学的物質,biological,細胞株,Cell line,細胞株を入力してください,Please enter Cell line,NULL,2022-08-09 02:57:35.896387+00,生物学的物質/細胞株,biological/Cell line
0dde5969-3039-739b-b33b-97df40450790,b4ce4016-e2bf-e5a1-7cae-ed496c7a776f,sample.specific.biological.protein-name,生物学的物質,biological,タンパク名,Protein name,タンパク名を入力してください,Please enter Protein name,NULL,2022-08-09 02:57:35.896387+00,生物学的物質/タンパク名,biological/Protein name
0dde5969-3039-739b-b33b-97df40450790,8c9b1a88-1530-24d3-4b2e-5441eee5c24f,sample.specific.biological.gene-name,生物学的物質,biological,遺伝子名,Gene name,遺伝子名を入力してください,Please enter Gene name,NULL,2022-08-09 02:57:35.896387+00,生物学的物質/遺伝子名,biological/Gene name
0dde5969-3039-739b-b33b-97df40450790,047e30f3-f294-e58d-cbe4-6bb588bf4cf8,sample.specific.biological.ncbi-accession-number,生物学的物質,biological,NCBIアクセッション番号,NCBI accession number,NCBIアクセッション番号を入力してください,Please enter NCBI accession number,NULL,2022-08-09 02:57:35.896387+00,生物学的物質/NCBIアクセッション番号,biological/NCBI accession number
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,9270879d-d94e-4d3f-2d5c-19568e040004,sample.specific.organic-material.inchi,有機材料,organic material,InChI,InChI,InChIを入力してください,Please enter InChI,NULL,2022-08-01 09:32:47.462801+00,有機材料/InChI,organic material/InChI
01cb3c01-37a4-5a43-d8ca-f523ca99a75b,3edadcff-8a85-51d9-708f-8f76bf055377,sample.specific.organic-material.inchi-key,有機材料,organic material,InChI key,InChI key,InChI keyを入力してください,Please enter InChI key,NULL,2022-08-01 09:32:47.462801+00,有機材料/InChI key,organic material/InChI key
    """.strip()


def preparedb() -> sqlite3.connect:
    """中間DBとして使用しているSQLiteの使用準備をする機能"""
    conn1 = sqlite3.connect(":memory:")  # オンメモリ
    # db_name = './aaa.db'               # ファイルに出力する場合はこの2行の方に切り替える
    # conn1 = sqlite3.connect(db_name)
    cur1 = conn1.cursor()

    csv_data = [
        "csv_dict_term",
        "csv_sample_general_sample_term",
        "csv_sample_sample_class",
        "csv_sample_specific_sample_term",
    ]

    # 既にテーブルがある場合は削除
    for d in csv_data:
        tablename = d.replace("csv_", "")
        query = f"DROP TABLE IF EXISTS {tablename};"
        cur1.execute(query)

    # テーブル作成
    tablename = "dict_term"
    query = f"""
CREATE TABLE {tablename} (
     `id`                 TEXT
    ,`name_ja`            TEXT
    ,`name_en`            TEXT
    ,`hint_ja`            TEXT
    ,`hint_en`            TEXT
    ,`term_uri`           TEXT
    ,`created`            TEXT
    ,PRIMARY KEY (id)
);
    """
    cur1.execute(query)

    tablename = "sample_general_sample_term"
    query = f"""
CREATE TABLE {tablename} (
     `term_id`            TEXT
    ,`key_name`           TEXT
    ,PRIMARY KEY (term_id)
);
    """
    cur1.execute(query)

    tablename = "sample_sample_class"
    query = f"""
CREATE TABLE {tablename} (
     `id`                TEXT
    ,`name_ja`           TEXT
    ,`name_en`           TEXT
    ,PRIMARY KEY (id)
);
    """
    cur1.execute(query)

    tablename = "sample_specific_sample_term"
    query = f"""
CREATE TABLE {tablename} (
     `sample_class_id`    TEXT
    ,`term_id`            TEXT
    ,`key_name`           TEXT
    ,PRIMARY KEY (sample_class_id,term_id)
);
    """
    cur1.execute(query)

    # 一度確定
    conn1.commit()

    # 実データ投入

    tablename = "dict_term"
    query = (
        f"INSERT INTO {tablename} "
        "(`id`, `name_ja`, `name_en`, `hint_ja`, `hint_en`, `term_uri`, `created`) "
        "VALUES(?,?,?,?,?,?,?);"
    )
    csv_string = csv_dict_term()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    next(reader)  # 1行目はヘッダ行
    for row in reader:
        cur1.execute(query, row)

    tablename = "sample_general_sample_term"
    query = f"INSERT INTO {tablename} " "(`term_id`,`key_name`) " "VALUES(?,?);"
    csv_string = csv_sample_general_sample_term()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    next(reader)  # 1行目はヘッダ行
    for row in reader:
        cur1.execute(query, row[:2])

    tablename = "sample_sample_class"
    query = f"INSERT INTO {tablename} " "(`id`,`name_ja`,`name_en`) " "VALUES(?,?,?);"
    csv_string = csv_sample_sample_class()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    next(reader)  # 1行目はヘッダ行
    for row in reader:
        cur1.execute(query, row[:3])

    tablename = "sample_specific_sample_term"
    query = (
        f"INSERT INTO {tablename} "
        "(`sample_class_id`,`term_id`,`key_name`) "
        "VALUES(?,?,?);"
    )
    csv_string = csv_sample_specific_sample_term()
    csvfile = io.StringIO(csv_string)
    reader = csv.reader(csvfile, skipinitialspace=True)
    next(reader)  # 1行目はヘッダ行
    for row in reader:
        cur1.execute(query, row[:3])

    # 確定
    conn1.commit()

    return conn1


def writable_dir(dir: str) -> str:
    """指定されたフォルダが書き込み可能かどうかのチェックする機能"""
    if os.access(dir, os.W_OK) and os.path.isdir(dir):
        return os.path.abspath(dir)
    else:
        raise argparse.ArgumentTypeError(dir + " is not writable or does not exist.")


def prepare() -> argparse.Namespace:
    """引数の処理をする機能"""
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dir",
        type=str,
        default="./template",
        # required=True,
        help="Specify foldername where template files are stored[default=./template]",
    )
    parser.add_argument(
        "--out",
        # type=str,
        type=writable_dir,
        default="./",
        # required=True,
        help="Specify foldername where Excel file will be output [default=./]",
    )
    parser.add_argument(
        "-v", "--verbosity", action="count", default=0, help="Increase output verbosity"
    )
    args = parser.parse_args()

    # Templateファイル群の存在チェック
    p_dir = Path(args.dir)
    if not p_dir.exists():
        print(f'"{args.dir}" is not found.', file=sys.stderr)
        sys.exit(1)
    # JSON files
    for json in ["metadata-def.json", "catalog.schema.json", "invoice.schema.json"]:
        p_file = p_dir.joinpath(json)
        if not p_file.exists():
            print(f'"{p_file}" is not found in "{args.dir}".', file=sys.stderr)
            sys.exit(1)

    # OK
    return args


def main(args: argparse.Namespace) -> None:
    """各種Excelシートを作成する機能"""
    wb = Workbook()

    # テンプレートとなるシート作成
    template_sheet(wb)

    conn = preparedb()

    # テンプレートファイル読み込み
    template_dir = args.dir
    if args.verbosity >= 2:
        print(f"template files dir: {template_dir}")
    # templates = parse_json('./template')
    templates = parse_json(template_dir)

    # シートの生成
    # explanation_sheet(wb, templates.get('catalog.schema.json', {}))
    explanation_sheet(wb, templates)
    matadata_def_sheet(wb, templates.get("metadata-def.json", {}))
    catalog_schema_sheet(wb, templates.get("catalog.schema.json", {}))
    invoice_schema_sheet(wb, templates.get("invoice.schema.json", {}), conn)
    dict_term_sheet(wb)
    sample_general_sample_term_sheet(wb)
    sample_sample_class_sheet(wb)
    sample_specific_sample_term_sheet(wb)

    # 不要シートの削除
    rm_template_sheet(wb)

    # 全シートへのフォーカスの解除
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = False

    # 開いたときに表示されるシートを設定
    # wb.active = wb.sheetnames.index('要件定義(catalog.schema.json)')
    # wb.active = wb.sheetnames.index('要件定義(invoice.schema.json)')

    # 出力ファイル名の取得
    output_dir = args.out
    # template_dir = args.dir
    prefix = "RDEDatasetTemplateSheet"
    output_file = (
        f'{Path(f"{output_dir}")}/{prefix}_{Path(f"{template_dir}").name}.xlsx'
    )
    if args.verbosity >= 2:
        print(f"output_file: {output_file}")

    # Excelファイルの保存
    wb.save(output_file)

    conn.close()


if __name__ == "__main__":
    args = prepare()
    main(args)
